import bossmaster
from filtering import (
    _calc_edu_bonus,
    _keyword_found,
    _parse_candidate_salary_range,
    check_required_condition,
    filter_candidate,
    parse_experience_years,
)
from storage import load_candidates_all, save_candidates_all
from doc_parser import _extract_salary_range
from constants import SCORE_THRESHOLD_STRONG
import contextlib
import io
import json
import os
import tempfile
from pathlib import Path
from unittest.mock import patch


def test_parse_experience_years_supports_arabic_and_chinese_numbers():
    cases = {
        "3 年经验": 3,
        "三年以上 Java 经验": 3,
        "十二年开发经验": 12,
        "两年工作经验": 2,
        "没有年限": None,
    }
    for text, expected in cases.items():
        assert parse_experience_years(text) == expected, text


def test_bossmaster_keeps_filtering_compatibility_exports():
    assert bossmaster.filter_candidate is filter_candidate
    assert bossmaster.check_required_condition is check_required_condition
    assert bossmaster.parse_experience_years is parse_experience_years


def test_bossmaster_keeps_storage_compatibility_exports():
    assert bossmaster.load_candidates_all is load_candidates_all
    assert bossmaster.save_candidates_all is save_candidates_all


def test_extract_job_salary_range_handles_numeric_and_negotiable_text():
    assert _extract_salary_range("薪资范围：12k-15k") == (12, 15)
    assert _extract_salary_range("月薪: 20K-30K") == (20, 30)
    assert _extract_salary_range("薪资面议") == (None, None)
    assert _extract_salary_range("薪资可谈") == (None, None)


def test_parse_candidate_salary_range_from_summary_first_line():
    assert _parse_candidate_salary_range("15-16K\n统招本科，5 年 Java") == (15, 16)
    assert _parse_candidate_salary_range("20-35K·15薪\n本科") == (20, 35)
    assert _parse_candidate_salary_range("18K\n本科") == (18, 18)
    assert _parse_candidate_salary_range("面议\n本科") == (None, None)


def test_keyword_matching_uses_word_boundaries_for_english_terms():
    assert _keyword_found("AI Agent and LLM platform", "AI") is True
    assert _keyword_found("email platform", "AI") is False
    assert _keyword_found("熟悉智能体和知识库", "智能体") is True


def test_education_bonus_tiers_are_stable():
    assert _calc_edu_bonus("博士学历") == 10
    assert _calc_edu_bonus("211 硕士") == 9
    assert _calc_edu_bonus("硕士") == 7
    assert _calc_edu_bonus("985 本科") == 6
    assert _calc_edu_bonus("统招本科") == 3


def test_required_conditions_support_string_or_and():
    assert check_required_condition("统招本科，5 年 Java", "统招本科")["passed"] is True
    risky_result = check_required_condition("成教本科，5 年 Java", "统招本科")
    assert risky_result["passed"] is True
    assert "学历形式待确认：疑似非统招本科" in risky_result["risk_flags"]

    workflow = {"type": "or", "items": ["activiti", "camunda", "flowable"]}
    assert check_required_condition("有 Camunda 项目经验", workflow)["passed"] is True
    assert check_required_condition("只有 Spring Boot 经验", workflow)["passed"] is False

    stack = {"type": "and", "items": ["Java", "MySQL", "Redis"]}
    assert check_required_condition("Java MySQL Redis", stack)["passed"] is True
    assert check_required_condition("Java MySQL", stack)["passed"] is False


def test_filter_candidate_scores_and_hard_rejections_are_stable():
    rule = {
        "min_exp": 4,
        "edu": "本科",
        "work_location": "南京",
        "salary_min": 12,
        "salary_max": 15,
        "required_conditions": ["统招本科"],
        "keywords": [
            {"name": "Java", "weight": 2},
            {"name": "Spring Cloud", "weight": 2},
            {"name": "MySQL", "weight": 1},
            {"name": "Redis", "weight": 1},
        ],
    }

    passed, score, details = filter_candidate(
        "15-16K\n南京，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is True
    assert score >= SCORE_THRESHOLD_STRONG
    assert details["skill_matched_count"] == 4

    passed, _, details = filter_candidate(
        "18-22K\n南京，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is False
    assert "薪资不匹配" in details["reason"]

    passed, _, details = filter_candidate(
        "15-16K\n上海，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is False
    assert "地点不符" in details["reason"]


def test_geek_card_api_payload_builds_complete_candidate_summary():
    payload = {
        "zpData": {
            "geekList": [
                {
                    "encryptGeekId": "encrypted-g-api-1",
                    "geekCard": {
                        "geekId": 123456,
                        "encGeekId": "encrypted-g-api-1",
                        "geekName": "张三",
                        "ageDesc": "32岁",
                        "geekDegree": "本科",
                        "geekWorkYear": "8年",
                        "expectLocationName": "南京",
                        "salary": "15-20K",
                        "expectPositionName": "Python 后端工程师",
                        "geekDesc": {"content": "熟悉金融数据平台"},
                        "geekEdus": [
                            {
                                "school": "南京大学",
                                "major": "计算机科学",
                                "degreeName": "本科",
                                "startDate": "2008",
                                "endDate": "2012",
                            }
                        ],
                        "geekWorks": [
                            {
                                "company": "某证券公司",
                                "positionName": "高级开发工程师",
                                "responsibility": "负责 ETL 调度、Python 数据分析和 Oracle 数据库开发",
                                "workEmphasisList": ["Python", "ETL", "Oracle"],
                                "startDate": "2018",
                                "endDate": "至今",
                            }
                        ],
                    }
                }
            ]
        }
    }

    candidates = bossmaster._extract_candidates_from_api_payload(payload)

    assert candidates == [
        {
            "geek_id": "encrypted-g-api-1",
            "name": "张三",
            "summary": candidates[0]["summary"],
            "structured": candidates[0]["structured"],
            "_api_profile": candidates[0]["_api_profile"],
        }
    ]
    summary = candidates[0]["summary"]
    structured = candidates[0]["structured"]
    api_profile = candidates[0]["_api_profile"]
    assert structured.get('exp_years') == 8
    assert structured.get('age') == 32
    assert structured.get('degree') == "本科"
    assert structured.get('city') == "南京"
    # _api_profile 结构化画像
    assert api_profile['personal_summary'] == "熟悉金融数据平台"
    assert len(api_profile['educations']) == 1
    assert api_profile['educations'][0]['school'] == "南京大学"
    assert len(api_profile['works']) == 1
    assert api_profile['works'][0]['company'] == "某证券公司"
    assert api_profile['works'][0]['skills'] == ["Python", "ETL", "Oracle"]
    assert "工作职责：负责 ETL 调度、Python 数据分析和 Oracle 数据库开发" in summary
    assert "技能标签：Python、ETL、Oracle" in summary
    assert "教育经历：南京大学 计算机科学 本科 2008 2012" in summary


def test_api_candidate_summary_participates_in_existing_filtering():
    geek_card = {
        "geekId": "g-api-2",
        "geekName": "李四",
        "ageDesc": "30岁",
        "geekDegree": "本科",
        "geekWorkYear": "6年",
        "expectLocationName": "南京",
        "salary": "14-16K",
        "geekWorks": [
            {
                "positionName": "数据开发工程师",
                "responsibility": "负责 Python 爬虫、SQL 数据处理和 Agent 工作流开发",
                "workEmphasisList": [{"name": "Python"}, {"name": "SQL"}, {"name": "Agent"}],
            }
        ],
    }
    rule = {
        "min_exp": 5,
        "edu": "本科",
        "work_location": "南京",
        "salary_min": 12,
        "salary_max": 16,
        "keywords": ["Python", "SQL", "Agent"],
    }

    summary = bossmaster._build_candidate_summary_from_geek_card(geek_card)
    passed, score, details = filter_candidate(summary, rule)

    assert passed is True
    assert score >= SCORE_THRESHOLD_STRONG
    assert details["skill_matched_count"] == 3


def test_dom_scan_uses_conservative_empty_limit_without_api_listener():
    first_dom_batch = [
        {
            "geek_id": f"g-dom-{i}",
            "name": f"候选人{i}",
            "text": f"本科，{i + 4}年 Java 开发工程师",
        }
        for i in range(15)
    ]
    dom_batches = [first_dom_batch, [], [], [], [], []]

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, *_args, **_kwargs):
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None) as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=([], "")) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', side_effect=dom_batches) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=6)

    assert len(candidates) == 15
    assert mock_dom_extract.call_count == 6
    # 无法从当前页构造 API 分页时，降级到 listener + DOM，但不刷新页面。
    mock_start_listener.assert_called_once()
    assert page.refresh_count == 0
    # API 消费调用：每轮 1 次，共 6 轮
    assert mock_consume_api.call_count == 6


def test_recommend_api_pagination_builds_from_current_iframe_jobid():
    class FakeFrame:
        def run_js(self, script):
            assert script == 'return location.href'
            return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"

    pagination = bossmaster._build_recommend_api_pagination_from_page(FakeFrame())

    assert pagination["base_url"] == "https://www.zhipin.com/wapi/zpjob/rec/geek/list"
    assert pagination["page_param"] == "page"
    assert pagination["page_size"] is None
    assert pagination["query_params"]["jobId"] == "job-123"
    assert pagination["query_params"]["page"] == "1"


def test_scan_uses_direct_api_pagination_without_refresh_or_dom_scroll():
    class FakeFrame:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakeFrame()
    api_pages = [
        ([{"geek_id": "g-api-1", "name": "张三", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-api-2", "name": "李四", "summary": "本科，6年 Java", "structured": {"exp_years": 6}}], False),
        ([], False),
    ]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener') as mock_start_listener, \
            patch('bossmaster._fetch_api_page_result', side_effect=api_pages) as mock_fetch, \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch') as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=3)

    assert [c["geek_id"] for c in candidates] == ["g-api-1", "g-api-2"]
    assert page.refresh_count == 0
    assert mock_fetch.call_count == 3
    mock_start_listener.assert_not_called()
    mock_consume_api.assert_not_called()
    mock_dom_extract.assert_not_called()


def test_scan_falls_back_to_refresh_listener_when_direct_api_unavailable():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            if "document.body" in script:
                return "Java 工程师 _ 南京 15-20K"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=FakeListener()) as mock_start_listener, \
            patch('bossmaster._fetch_api_page_result', return_value=([], False)) as mock_fetch, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=(
                [{"geek_id": "g-api-refresh", "name": "王五", "summary": "本科，7年 Java", "structured": {"exp_years": 7}}],
                "https://www.zhipin.com/wapi/zpjob/rec/geek/list",
            )) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch') as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=1)

    assert [c["geek_id"] for c in candidates] == ["g-api-refresh"]
    assert page.refresh_count == 1
    mock_start_listener.assert_called_once()
    mock_fetch.assert_called_once()
    mock_consume_api.assert_called_once()
    mock_dom_extract.assert_not_called()


def test_scan_stops_on_api_risk_status_without_refresh_or_dom_fallback():
    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._fetch_api_page_result', side_effect=bossmaster.ApiRiskBlocked(429, 1)) as mock_fetch, \
            patch('bossmaster._start_recommend_api_listener') as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch') as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=1)

    assert candidates == []
    assert page.refresh_count == 0
    mock_fetch.assert_called_once()
    mock_start_listener.assert_not_called()
    mock_consume_api.assert_not_called()
    mock_dom_extract.assert_not_called()


def test_collect_captcha_diagnostic_writes_json_without_screenshot():
    class FakePage:
        def run_js(self, script):
            if script == "return location.href":
                return "https://www.zhipin.com/web/chat/recommend"
            if script == "return document.title":
                return "推荐牛人"
            if "document.body" in script:
                return "请完成安全验证"
            return ""

    with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(bossmaster, "BASE_DIR", Path(tmpdir)), \
            patch.object(bossmaster, "get_iframe", return_value=None):
        path = bossmaster._collect_captcha_diagnostic(
            FakePage(),
            detail="主页面检测到安全验证弹窗",
            stage="scan",
        )

        assert path is not None
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert payload["stage"] == "scan"
        assert payload["url"] == "https://www.zhipin.com/web/chat/recommend"
        assert "安全验证" in payload["visible_text_excerpt"]


def test_refresh_listener_stops_when_refresh_changes_job_identity():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                job_id = "job-after" if self.refresh_count else "job-before"
                return f"https://www.zhipin.com/web/frame/recommend/?jobid={job_id}&status=0"
            if "document.body" in script:
                return "默认岗位 _ 南京 10-15K" if self.refresh_count else "目标岗位 _ 南京 15-20K"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=FakeListener()), \
            patch('bossmaster._fetch_api_page_result', return_value=([], None)), \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch') as mock_dom_extract:
        output = io.StringIO()
        with contextlib.redirect_stdout(output):
            candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=1)

    assert candidates == []
    assert page.refresh_count == 1
    assert "刷新后岗位已变化，请先将目标岗位设为默认岗位" in output.getvalue()
    mock_consume_api.assert_not_called()
    mock_dom_extract.assert_not_called()


def test_find_card_by_scroll_returns_to_top_after_current_position_miss():
    class FakeTarget:
        def __init__(self):
            self.scroll = 5000
            self.scripts = []

        def run_js(self, script):
            self.scripts.append(script)
            if "window.scrollTo(0, 0)" in script:
                self.scroll = 0
                return None
            if "scrollTop" in script and "return" in script:
                return self.scroll
            return None

        def ele(self, *_args, **_kwargs):
            return "card" if self.scroll == 0 else None

    target = FakeTarget()

    with patch('bossmaster.time.sleep'), patch('bossmaster._human_delay', return_value=0):
        card = bossmaster._find_card_by_scroll(target, 'css:[data-geekid="g1"]')

    assert card == "card"
    assert any("window.scrollTo(0, 0)" in script for script in target.scripts)


def test_export_to_excel_keeps_full_candidate_summary_in_detail_column():
    long_summary = (
        "15-18K\n南京，统招本科，6 年 Python 经验\n"
        + "工作职责：负责数据仓库建设、ETL 调度、SQL 优化和业务指标分析。"
        + "技能标签：Python、SQL、ETL、Oracle。"
        + "项目说明：" + "A" * 260
    )
    with tempfile.TemporaryDirectory() as tmpdir:
        output = os.path.join(tmpdir, "candidates.xlsx")
        bossmaster.export_to_excel(
            [
                {
                    "geek_id": "g-excel-1",
                    "name": "王五",
                    "summary": long_summary,
                    "job_name": "数据分析师",
                    "match_score": 80,
                    "recommend_level": "强烈推荐",
                    "greet_sent": False,
                    "manual_review_required": True,
                    "risk_flags": ["学历形式待确认：疑似非统招本科"],
                    "auto_greet_blocked_reason": "学历形式待确认",
                }
            ],
            output,
        )

        from openpyxl import load_workbook

        workbook = load_workbook(output)
        assert workbook.sheetnames == ["全部候选人", "数据分析师", "统计摘要"]

        sheet = workbook["全部候选人"]
        headers = [cell.value for cell in sheet[1]]
        detail_col = headers.index("详细信息") + 1
        manual_review_col = headers.index("是否需人工确认") + 1
        risk_col = headers.index("风险提示") + 1

        assert sheet.cell(row=2, column=detail_col).value == long_summary
        assert sheet.cell(row=2, column=manual_review_col).value == "是"
        assert sheet.cell(row=2, column=risk_col).value == "学历形式待确认：疑似非统招本科"

        job_sheet = workbook["数据分析师"]
        assert job_sheet.cell(row=2, column=1).value == 1

        summary_sheet = workbook["统计摘要"]
        summary_headers = [cell.value for cell in summary_sheet[1]]
        total_col = summary_headers.index("总人数") + 1
        avg_col = summary_headers.index("平均分") + 1
        assert summary_sheet.cell(row=2, column=total_col).value == 1
        assert summary_sheet.cell(row=2, column=avg_col).value == "80.0"


def test_auto_greet_skips_manual_review_candidates():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-risk",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {
            "min_exp": 0,
            "edu": "本科",
            "required_conditions": ["统招本科"],
            "keywords": ["Java"],
        },
    }
    raw_candidates = [{
        "geek_id": "g-risk-1",
        "name": "赵六",
        "summary": "20K\n北京，专升本，5 年 Java 开发",
    }]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "send_greeting_on_list_page") as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert len(result) == 1
    assert result[0]["manual_review_required"] is True
    assert result[0]["greet_sent"] is False
    mock_greet.assert_not_called()


def test_auto_greet_uses_page_order_not_score_order():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-order",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": "g-page-first", "name": "张三", "summary": "本科，3 年 Java"},
        {"geek_id": "g-page-second", "name": "李四", "summary": "本科，10 年 Java"},
    ]
    filter_results = [
        (True, 65, {"skill_matches": ["Java"]}),
        (True, 95, {"skill_matches": ["Java"]}),
    ]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", side_effect=filter_results), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "send_greeting_on_list_page", return_value=(True, "成功")) as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert [call.args[1] for call in mock_greet.call_args_list] == [
        "g-page-first",
        "g-page-second",
    ]


def test_auto_greet_limit_triggers_notice_and_caps_greetings():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-limit",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": f"g-{i}", "name": f"候选人{i}", "summary": "本科，5 年 Java"}
        for i in range(25)
    ]
    notices = []

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "send_greeting_on_list_page", return_value=(True, "成功")) as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
            notice_callback=lambda title, message: notices.append((title, message)),
        )

    assert mock_greet.call_count == bossmaster.AUTO_GREET_RUN_LIMIT
    assert notices
    assert "剩余 5 人下次继续" in notices[0][1]
    assert "再次运行同一岗位扫描" in notices[0][1]


def test_filter_candidate_age_boundaries_are_stable():
    rule = {
        "min_exp": 0,
        "edu": "不限",
        "max_age": 35,
        "keywords": ["Java"],
    }

    passed, _, _ = filter_candidate("35岁，Java 开发", rule)
    assert passed is True

    passed, _, details = filter_candidate("年龄：36 岁，Java 开发", rule)
    assert passed is False
    assert "年龄不符" in details["reason"]


def test_filter_candidate_flags_non_regular_bachelor_even_with_school_mark():
    rule = {
        "min_exp": 0,
        "edu": "本科",
        "required_conditions": ["统招本科"],
        "keywords": ["Java"],
    }

    passed, _, details = filter_candidate("985 本科，专升本，5 年 Java", rule)
    assert passed is True
    assert details["manual_review_required"] is True
    assert "学历形式待确认：疑似非统招本科" in details["risk_flags"]

    passed, _, _ = filter_candidate("全日制本科，5 年 Java", rule)
    assert passed is True


def test_save_candidates_all_deduplicates_by_geek_id_and_job_name():
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            with contextlib.redirect_stdout(io.StringIO()):
                save_candidates_all([
                    {
                        "geek_id": "g1",
                        "job_name": "Java",
                        "match_score": 70,
                        "greet_sent": True,
                        "greeting_in_progress": True,
                    },
                    {
                        "geek_id": "g1",
                        "job_name": "Java",
                        "match_score": 80,
                        "greet_sent": False,
                    },
                    {
                        "geek_id": "g1",
                        "job_name": "Python",
                        "match_score": 60,
                        "greet_sent": False,
                    },
                ])

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                saved = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert len(saved) == 2
    java = next(c for c in saved if c["job_name"] == "Java")
    python = next(c for c in saved if c["job_name"] == "Python")
    assert java["match_score"] == 80
    assert java["greet_sent"] is True
    assert "greeting_in_progress" not in java
    assert python["match_score"] == 60


def test_save_candidates_all_filters_below_55():
    """低于 55 分的候选人不应写入 candidates_all.json"""
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            with contextlib.redirect_stdout(io.StringIO()):
                save_candidates_all([
                    {"geek_id": "g1", "job_name": "Java", "match_score": 80},
                    {"geek_id": "g2", "job_name": "Java", "match_score": 55},
                    {"geek_id": "g3", "job_name": "Java", "match_score": 54},
                    {"geek_id": "g4", "job_name": "Java", "match_score": 30},
                ])

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                saved = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert len(saved) == 2
    ids = {c["geek_id"] for c in saved}
    assert ids == {"g1", "g2"}


def test_load_candidates_all_restores_from_backup_when_main_json_is_corrupt():
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            backup_data = [{"geek_id": "g1", "job_name": "Java", "greet_sent": True}]
            with open("candidates_all.json", "w", encoding="utf-8") as f:
                f.write("{broken json")
            with open("candidates_all.json.bak", "w", encoding="utf-8") as f:
                json.dump(backup_data, f, ensure_ascii=False)

            with contextlib.redirect_stdout(io.StringIO()):
                loaded = load_candidates_all()

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                restored = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert loaded == backup_data
    assert restored == backup_data


def test_save_candidates_all_accepts_explicit_path():
    with tempfile.TemporaryDirectory() as tmpdir:
        target = os.path.join(tmpdir, "nested_candidates.json")
        with contextlib.redirect_stdout(io.StringIO()):
            save_candidates_all([
                {"geek_id": "g1", "job_name": "Java", "match_score": 70},
            ], target)

        with open(target, "r", encoding="utf-8") as f:
            saved = json.load(f)

    assert saved == [{"geek_id": "g1", "job_name": "Java", "match_score": 70}]


# ========== load_job_config ==========

def test_load_job_config_jobs_key_format():
    """支持 "jobs" 键格式的配置文件。"""
    config = {
        "jobs": {
            "Java工程师": {
                "min_exp": 3,
                "edu": "本科",
                "keywords": ["Java", "Spring"]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, default = load_job_config()
        assert "Java工程师" in jobs
        assert default is None
        assert jobs["Java工程师"]["min_exp"] == 3
    finally:
        os.unlink(tmp_path)


def test_load_job_config_extracts_default_rule():
    """default 规则应从 job_requirements 中提取出来单独返回。"""
    config = {
        "job_requirements": {
            "default": {"min_exp": 0, "edu": "不限", "keywords": []},
            "Python工程师": {"min_exp": 2, "keywords": ["Python"]},
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, default = load_job_config()
        assert "default" not in jobs
        assert default is not None
        assert default["edu"] == "不限"
        assert "Python工程师" in jobs
    finally:
        os.unlink(tmp_path)


def test_load_job_config_strips_spaces_from_job_name():
    """岗位名称中的空格应被移除。"""
    config = {"jobs": {"Java 工程师": {"keywords": ["Java"]}}}
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        assert "Java工程师" in jobs
        assert "Java 工程师" not in jobs
    finally:
        os.unlink(tmp_path)


def test_load_job_config_deduplicates_keywords_case_insensitive():
    """关键词应按小写去重，保留首次出现的格式。"""
    config = {
        "jobs": {
            "Dev": {
                "keywords": ["Java", "java", "JAVA", "Spring"]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        kws = jobs["Dev"]["keywords"]
        assert len(kws) == 2  # Java (case-insensitive dedup) + Spring
        assert kws[0] == "Java"  # 保留首次出现
        assert "Spring" in kws
    finally:
        os.unlink(tmp_path)


def test_load_job_config_deduplicates_dict_format_keywords():
    """dict 格式关键词也应按 name 去重。"""
    config = {
        "jobs": {
            "Dev": {
                "keywords": [
                    {"name": "Java", "weight": 2},
                    {"name": "java", "weight": 1},
                    {"name": "Spring", "weight": 1},
                ]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        kws = jobs["Dev"]["keywords"]
        assert len(kws) == 2
        assert kws[0]["name"] == "Java"
        assert kws[0]["weight"] == 2  # 保留首次出现的 weight
    finally:
        os.unlink(tmp_path)


def test_load_job_config_missing_file_returns_default():
    """配置文件不存在时返回默认配置。"""
    with patch('bossmaster.CONFIG_PATH', '/nonexistent/path/job_config.json'):
        from bossmaster import load_job_config
        with contextlib.redirect_stdout(io.StringIO()):
            jobs, default = load_job_config()
        assert default is None
        assert "default" in jobs
        assert jobs["default"]["edu"] == "不限"


def test_load_job_config_corrupt_json_returns_default():
    """配置文件 JSON 损坏时返回默认配置。"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        f.write("{broken json")
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            with contextlib.redirect_stdout(io.StringIO()):
                jobs, default = load_job_config()
        assert default is None
        assert "default" in jobs
    finally:
        os.unlink(tmp_path)


# ========== extract_summary_info ==========

def test_extract_summary_info_full_text():
    """完整摘要文本应提取所有字段。"""
    from bossmaster import extract_summary_info
    text = "15-20K\n30 岁，6 年经验，本科\n离职-某某科技有限公司\n南京\n熟悉 Java、Spring、MySQL、Redis"
    info = extract_summary_info(text)
    assert info['salary'] == '15-20K'
    assert info['age'] == '30'
    assert info['exp_years'] == '6'
    assert info['education'] == '本科'
    assert info['job_status'] == '离职'
    assert '某某科技' in info['company']
    assert 'Java' in info['skills']
    assert 'MySQL' in info['skills']


def test_extract_summary_info_negotiable_salary():
    """面议薪资应正确识别。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("面议\n本科，3 年经验")
    assert info['salary'] == '面议'


def test_extract_summary_info_empty_text():
    """空文本应返回全空字典。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("")
    assert all(v == '' for v in info.values())


def test_extract_summary_info_education_priority():
    """学历应取最高级别（博士 > 硕士 > 本科）。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("本科，硕士在读，博士毕业")
    assert info['education'] == '博士'


def test_extract_summary_info_status_with_company():
    """在职/离职状态和公司名提取。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("在职-阿里巴巴集团")
    assert info['job_status'] == '在职'
    assert info['company'] == '阿里巴巴集团'
