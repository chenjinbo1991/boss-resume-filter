"""
BOSS 直聘候选人智能提取工具
支持 Excel 导出
"""
from __future__ import annotations

import json
import logging
import os
import random
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse

from DrissionPage import ChromiumPage

from constants import (
    SCORE_THRESHOLD_PASS,
    SCORE_THRESHOLD_RECOMMEND,
    SCORE_THRESHOLD_STRONG,
    SCROLL_PX,
    MAX_SCROLL_SEARCH,
    MAX_ROUNDS_DEFAULT,
    EMPTY_ROUNDS_LIMIT,
    GREET_FAIL_LIMIT,
    GREET_UNCERTAIN_LIMIT,
    CAPTCHA_MAX_WAIT,
    CAPTCHA_CHECK_INTERVAL,
    API_PAGE_DELAY_CENTER,
    API_PAGE_DELAY_SPREAD,
    API_CANDIDATE_LIMIT_DEFAULT,
    AUTO_GREET_RUN_LIMIT,
    GREET_DELAY_CENTER,
    GREET_DELAY_SPREAD,
    GREET_BATCH_SIZE,
    GREET_BATCH_PAUSE_CENTER,
    GREET_BATCH_PAUSE_SPREAD,
)
from filtering import (
    _calc_edu_bonus,
    _extract_city,
    _keyword_found,
    _parse_candidate_salary_range,
    check_required_condition,
    evaluate_candidate,
    filter_candidate,
    parse_experience_years,
)
from paths import BASE_DIR, SELECTORS_PATH, CONFIG_PATH, CANDIDATES_PATH, CANDIDATES_XLSX_PATH
from storage import (
    build_blacklist_index,
    build_greeted_index,
    get_greeted_geek_ids,
    is_already_greeted,
    load_candidates_all,
    merge_candidates_all,
    persist_candidate_greeting_pending,
    persist_candidate_greeted,
    save_candidates_all,
)

logger = logging.getLogger(__name__)


def _read_app_version() -> str:
    """Read GUI version directly from gui_main module."""
    try:
        import gui_main
        return getattr(gui_main, '__version__', 'unknown')
    except Exception:
        return "unknown"


class StopRequested(Exception):
    """停止请求异常 — 用于立即终止扫描流程"""
    pass


class ApiRiskBlocked(Exception):
    """BOSS API 返回疑似风控状态码时立即熔断，不继续刷新或 DOM 滚动。"""

    def __init__(self, status: int | str, page_num: int):
        self.status = status
        self.page_num = page_num
        super().__init__(f"API page {page_num} returned risk status {status}")


RECOMMEND_PAGE_URL_PARTS = (
    "zhipin.com/web/chat/recommend",
    "zhipin.com/web/frame/recommend",
)
RECOMMEND_PAGE_URL = "https://www.zhipin.com/web/chat/recommend"
GREET_CONTEXT_VERSION = 1
GREET_CONTEXT_CAPTURE_LIMIT = 30
GREET_CONTEXT_MIN_SCORE = SCORE_THRESHOLD_PASS
GREET_CONTEXT_DELAY_CENTER = 1.2
GREET_CONTEXT_DELAY_SPREAD = 0.7
GREET_CONTEXT_BATCH_SIZE = 10
GREET_CONTEXT_BATCH_PAUSE_CENTER = 5.0
GREET_CONTEXT_BATCH_PAUSE_SPREAD = 2.5


def _ensure_recommend_page(page: Any, notice_callback=None, context: str = "运行") -> bool:
    """确认当前仍在推荐牛人页面；页面跑偏时通知 GUI 并让调用方停止本轮。"""
    error_detail = "无法读取当前页面 URL"
    try:
        current_url = str(getattr(page, "url", "") or "")
    except Exception as e:
        current_url = ""
        error_detail = f"无法读取当前页面 URL：{e}"

    if not current_url:
        try:
            current_url = str(page.run_js('return location.href') or "")
        except Exception as e:
            if error_detail == "无法读取当前页面 URL":
                error_detail = f"无法读取当前页面 URL：{e}"

    if current_url and any(part in current_url.lower() for part in RECOMMEND_PAGE_URL_PARTS):
        return True

    page_detail = current_url if current_url else error_detail

    message = (
        f"{context}时发现浏览器已经不在 BOSS 直聘推荐牛人页面。\n\n"
        f"当前页面：{page_detail}\n\n"
        f"请先在浏览器中切回推荐牛人页面：\n{RECOMMEND_PAGE_URL}\n\n"
        "本轮运行已停止，避免在错误页面继续刷新或自动操作。"
    )
    print(f"[WARN] 当前页面不是 BOSS 直聘推荐牛人页面，已停止本轮运行。当前页面：{page_detail}")
    if notice_callback:
        try:
            notice_callback("请切回推荐牛人页面", message)
        except Exception:
            pass
    return False


def _save_progress_on_exit() -> None:
    """加载磁盘数据并导出 Excel，供异常/中断 handler 统一调用。"""
    existing_all = load_candidates_all()
    if export_to_excel(existing_all, CANDIDATES_XLSX_PATH):
        print(f"[SAVE] Excel 文件：{CANDIDATES_XLSX_PATH.name}")
    print(f"已保存 {len(existing_all)} 个候选人的状态")


def _human_delay(center: float, spread: float = 0.3) -> float:
    """模拟人类操作延迟，在 center ± spread/2 范围内随机抖动，降低行为指纹风险"""
    return center + random.uniform(-spread / 2, spread / 2)


try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


_SELECTORS_CACHE = None


def load_selectors() -> dict[str, Any]:
    """加载 selectors.json 选择器配置（首次调用后缓存）"""
    global _SELECTORS_CACHE
    if _SELECTORS_CACHE is not None:
        return _SELECTORS_CACHE
    try:
        with open(SELECTORS_PATH, "r", encoding="utf-8") as f:
            _SELECTORS_CACHE = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logger.warning("加载 selectors.json 失败：%s，使用内置默认值", e)
        _SELECTORS_CACHE = {}
    return _SELECTORS_CACHE


def _sel(group: str, key: str, default: str = "") -> str:
    """从 selectors.json 获取选择器值，找不到则返回 default"""
    s = load_selectors()
    return s.get(group, {}).get(key, default)


def load_job_config() -> tuple[dict[str, Any], dict[str, Any]]:
    """从外部配置文件加载职位要求（支持多岗位）
    返回：(job_requirements, default_rule)
        - job_requirements: 各岗位的配置（不包含 default）
        - default_rule: 默认过滤规则
    """
    default_rule = None

    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)

                # 支持新旧两种格式
                if "jobs" in config:
                    job_requirements = config["jobs"]
                else:
                    job_requirements = config.get("job_requirements", {})

                # 提取 default 作为默认规则（如果存在）
                if "default" in job_requirements:
                    default_rule = job_requirements.pop("default")

                # 处理岗位名称去空格 + keywords 去重
                new_job_requirements = {}
                for job_name, rule in job_requirements.items():
                    # 岗位名称去除空格
                    clean_job_name = job_name.replace(" ", "")

                    if isinstance(rule, dict) and rule.get('keywords'):
                        seen = set()
                        unique_keywords = []
                        for kw in rule['keywords']:
                            # 支持两种格式：字符串 或 {"name": xxx, "weight": xxx}
                            if isinstance(kw, dict):
                                kw_lower = kw.get('name', '').lower()
                            else:
                                kw_lower = kw.lower()
                            if kw_lower not in seen:
                                seen.add(kw_lower)
                                unique_keywords.append(kw)
                        rule['keywords'] = unique_keywords

                    new_job_requirements[clean_job_name] = rule

                job_requirements = new_job_requirements

                return job_requirements, default_rule
        except Exception as e:
            print(f"读取配置文件出错：{e}")

    print("未找到配置文件，使用默认配置...")
    return {
        "default": {
            "min_exp": 0,
            "edu": "不限",
            "keywords": []
        }
    }, None


JOB_RULES = load_job_config()


def _normalize_job_name_for_match(job_name: str) -> str:
    """Normalize job names for selection matching while preserving configured keys."""
    return re.sub(r'\s+', '', job_name or '').casefold()


def _resolve_job_name(requested_job: str, job_rules: dict[str, Any]) -> str | None:
    """Resolve a user-selected job name to the configured job key."""
    if requested_job in job_rules:
        return requested_job

    requested_normalized = _normalize_job_name_for_match(requested_job)
    matches = [
        job_name for job_name in job_rules
        if _normalize_job_name_for_match(job_name) == requested_normalized
    ]
    if len(matches) == 1:
        return matches[0]
    return None


def extract_summary_info(text: str) -> dict[str, Any]:
    """从候选人摘要中提取结构化信息

    同时支持 DOM 格式（"15-20K\\n30 岁，6 年经验"）
    和 API 标签格式（"期望薪资：15-25K\\n年龄：30\\n经验：5年"）
    """
    info: dict[str, Any] = {
        'salary': '',
        'age': '',
        'exp_years': '',
        'education': '',
        'job_status': '',
        'company': '',
        'city': '',
        'skills': ''
    }

    if not text:
        return info

    lines = text.split('\n')
    first_line = lines[0].strip() if lines else ''

    # ---- 薪资 ----
    # API 标签格式
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("期望薪资："):
            val = stripped[len("期望薪资："):].strip()
            if '面议' in val:
                info['salary'] = '面议'
            else:
                # 15K-25K / 15-20K / 15-25（K 可选）
                m = re.search(r'(\d+(?:\.\d+)?)\s*[Kk]?\s*[-~～\-]\s*(\d+(?:\.\d+)?)\s*[Kk]?', val)
                if m:
                    info['salary'] = f"{m.group(1)}-{m.group(2)}K"
                else:
                    # 15-25薪
                    m = re.search(r'(\d+)\s*薪\s*[-~～\-]\s*(\d+)\s*薪', val)
                    if m:
                        info['salary'] = f"{m.group(1)}-{m.group(2)}K"
                    else:
                        # X万-Y万
                        m = re.search(r'([\d.]+)\s*[万萬]\s*[-~～\-]\s*([\d.]+)\s*[万萬]', val)
                        if m:
                            lo = int(float(m.group(1)) * 10)
                            hi = int(float(m.group(2)) * 10)
                            info['salary'] = f"{lo}-{hi}K"
                        else:
                            # 15K / 15K以上 / 15薪
                            m = re.search(r'(\d+)\s*[Kk薪千]', val)
                            if m:
                                info['salary'] = m.group(1) + 'K'
                            elif '面议' in val:
                                info['salary'] = '面议'
            break

    # DOM 格式兜底
    if not info['salary']:
        if '面议' in first_line:
            info['salary'] = '面议'
        else:
            salary_match = re.search(r'(\d+(?:\.\d+)?)\s*[Kk]?\s*[-~～\-]\s*(\d+(?:\.\d+)?)\s*[Kk]?', first_line)
            if salary_match:
                info['salary'] = f"{salary_match.group(1)}-{salary_match.group(2)}K"
            else:
                salary_match = re.search(r'(\d+(?:\.\d+)?)[Kk千]', first_line)
                if salary_match:
                    info['salary'] = salary_match.group(1) + 'K'

    # ---- 年龄 ----
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("年龄："):
            m = re.search(r'(\d+)', stripped[len("年龄："):])
            if m:
                info['age'] = m.group(1)
            break
    if not info['age']:
        age_match = re.search(r'(\d+)\s*岁', text)
        if age_match:
            info['age'] = age_match.group(1)

    # ---- 工作经验年限 ----
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("经验："):
            val = stripped[len("经验："):].strip()
            m = re.search(r'(\d+)', val)
            if m:
                info['exp_years'] = m.group(1)
            break
    if not info['exp_years']:
        # DOM 格式：BOSS 直聘常见 "10年以上"、"9年"、"6年经验"、"3年工作"
        parsed_exp = parse_experience_years(text)
        if parsed_exp is not None:
            info['exp_years'] = str(parsed_exp)

    # ---- 学历 ----
    edu_keywords = ['博士', '硕士', '本科', '大专', '高中', '中专']
    for edu in edu_keywords:
        if edu in text:
            info['education'] = edu
            break

    # ---- 求职状态和公司 ----
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("求职状态："):
            val = stripped[len("求职状态："):].strip()
            m = re.search(r'(离职|在职|在校|应届)[-\s·—]*(.*)', val)
            if m:
                info['job_status'] = m.group(1)
                if m.group(2).strip():
                    info['company'] = m.group(2).strip()
            else:
                info['job_status'] = val
            break
    if not info['job_status']:
        status_match = re.search(r'(离职|在职|在校|应届)[-·—]([^\n]+)', text)
        if status_match:
            info['job_status'] = status_match.group(1)
            info['company'] = status_match.group(2).strip()

    # ---- 城市 ----
    info['city'] = _extract_city(text)

    # ---- 技能关键词 ----
    skill_keywords = ['Java', 'Python', 'MySQL', 'Oracle', 'Redis', 'Kafka',
                      'Spring', 'MyBatis', 'Dubbo', 'Vue', 'React', 'Linux',
                      'Docker', 'K8s', 'Kubernetes', 'AWS', 'Azure', 'Git']
    found_skills = [s for s in skill_keywords if s in text]
    info['skills'] = ', '.join(found_skills)

    return info


def export_to_excel(candidates: list[dict[str, Any]], filename: str) -> bool:
    """将候选人数据导出为 Excel - 增强版

    功能：
        - 按匹配分从高到低排序
        - 按岗位分工作表
        - 统计摘要工作表
        - 颜色标识推荐指数和打招呼状态
        - 自动筛选和冻结窗格
    """
    if not OPENPYXL_AVAILABLE:
        return False

    def _format_score_breakdown(c: dict[str, Any]) -> str:
        breakdown = c.get('score_breakdown') or {}
        if not breakdown:
            return ""
        parts = [
            f"基础{breakdown.get('base', 0)}",
            f"技能{breakdown.get('skill', 0)}",
            f"经验{breakdown.get('experience', 0)}",
            f"学历{breakdown.get('education', 0)}",
            f"优先{breakdown.get('preferred', 0)}",
        ]
        # AI 调整分（LLM 评估后追加）
        ai_adj = breakdown.get('ai_adjustment')
        if ai_adj is not None and ai_adj != 0:
            sign = "+" if ai_adj > 0 else ""
            parts.append(f"AI{sign}{ai_adj}")
        # 简历二次评估调整分
        resume_adj = breakdown.get('resume_adjustment')
        if resume_adj is not None and resume_adj != 0:
            sign = "+" if resume_adj > 0 else ""
            parts.append(f"简历{sign}{resume_adj}")
        parts.append(f"总分{breakdown.get('total', c.get('match_score', 0))}")
        return " / ".join(parts)

    def _format_score_explanation(c: dict[str, Any]) -> str:
        explanation = c.get('score_explanation') or []
        if isinstance(explanation, list):
            return "\n".join(str(item) for item in explanation if item)
        return str(explanation)

    def _format_keyword_evidence(c: dict[str, Any]) -> str:
        evidence_items = c.get('keyword_evidence') or []
        lines = []
        for item in evidence_items:
            if not isinstance(item, dict):
                continue
            name = item.get('name', '')
            evidence = item.get('evidence', '')
            if name and evidence:
                lines.append(f"{name}: {evidence}")
            elif name:
                lines.append(str(name))
        return "\n".join(lines)

    def _format_risk_flags(c: dict[str, Any]) -> str:
        risk_flags = c.get('risk_flags') or []
        if isinstance(risk_flags, list):
            return "\n".join(str(flag) for flag in risk_flags if flag)
        return str(risk_flags)

    def _format_edu_detail(c: dict[str, Any]) -> str:
        """学历明细：优先使用 _api_profile，fallback 到 summary 解析。"""
        api_profile = c.get('_api_profile')
        if api_profile and api_profile.get('educations'):
            lines = []
            for edu in api_profile['educations']:
                parts = [edu.get(k, '') for k in ('school', 'major', 'degree')]
                parts = [p for p in parts if p]
                start = edu.get('start', '')
                end = edu.get('end', '')
                if start or end:
                    parts.append(f"{start}-{end}")
                if parts:
                    lines.append(" ".join(parts))
            return "\n".join(lines)
        # Fallback: 从 summary 文本提取
        summary = c.get('summary', '')
        lines = []
        for sline in summary.split('\n'):
            sline = sline.strip()
            if sline.startswith("教育经历："):
                lines.append(sline[len("教育经历："):].strip())
        return "\n".join(lines)

    def _format_recent_company(c: dict[str, Any]) -> str:
        """最近公司：从 _api_profile 取第一段工作经历的公司名。"""
        api_profile = c.get('_api_profile')
        if api_profile and api_profile.get('works'):
            for work in api_profile['works']:
                company = work.get('company', '')
                if company:
                    return company
        # Fallback: 从 summary 文本提取
        summary = c.get('summary', '')
        for sline in summary.split('\n'):
            sline = sline.strip()
            if sline.startswith("工作经历："):
                val = sline[len("工作经历："):].strip()
                # 取第一段经历的公司名
                parts = val.split()
                if parts:
                    return parts[0]
        return ""

    def _format_skills(c: dict[str, Any], summary_info_skills: str) -> str:
        """技能：优先从 _api_profile 聚合工作经历技能标签，fallback 到 DOM 关键词匹配。"""
        api_profile = c.get('_api_profile')
        if api_profile and api_profile.get('works'):
            tags = []
            seen = set()
            for work in api_profile['works']:
                for tag in (work.get('skills') or []):
                    tag = tag.strip()
                    if tag and tag not in seen:
                        seen.add(tag)
                        tags.append(tag)
            if tags:
                return "、".join(tags)
        return summary_info_skills

    try:
        # 按匹配分从高到低排序
        visible_candidates = [c for c in candidates if not c.get('blacklisted')]
        sorted_candidates = sorted(visible_candidates, key=lambda x: x.get('match_score', 0), reverse=True)

        data = []
        for i, c in enumerate(sorted_candidates):
            score = c.get('match_score', 0)
            # 根据匹配分计算推荐指数
            if score >= SCORE_THRESHOLD_STRONG:
                recommend_level = "强烈推荐"
            elif score >= SCORE_THRESHOLD_RECOMMEND:
                recommend_level = "推荐"
            elif score >= SCORE_THRESHOLD_PASS:
                recommend_level = "待定"
            else:
                continue  # 低于通过分直接过滤，不进入导出

            summary_info = extract_summary_info(c.get('summary', ''))
            # API 结构化数据覆盖文本解析（与 smart_scan_candidates 一致）
            structured = c.get('structured') or {}
            if structured.get('salary_min') and structured.get('salary_max'):
                summary_info['salary'] = f"{structured['salary_min']}-{structured['salary_max']}K"
            elif structured.get('salary_min'):
                summary_info['salary'] = f"{structured['salary_min']}K"
            if structured.get('exp_years'):
                summary_info['exp_years'] = str(structured['exp_years'])
            if structured.get('age'):
                summary_info['age'] = str(structured['age'])
            if structured.get('city'):
                summary_info['city'] = structured['city']
            if structured.get('job_status'):
                summary_info['job_status'] = structured['job_status']
            row = {
                # ① 身份
                '序号': i + 1,
                '岗位': c.get('job_name', ''),
                '姓名': c.get('name', '未知'),
                'geek_id': c.get('geek_id', ''),
                # ② 画像
                '年龄': summary_info['age'],
                '工作年限': summary_info['exp_years'],
                '学历': summary_info['education'],
                '学历明细': _format_edu_detail(c),
                '薪资': summary_info['salary'],
                '求职状态': summary_info['job_status'],
                '城市': summary_info['city'],
                '最近公司': _format_recent_company(c),
                '技能': _format_skills(c, summary_info['skills']),
                # ③ 评估
                '匹配分': score,
                '推荐指数': recommend_level,
                '技能匹配': c.get('skill_match_ratio', ''),
                '评分拆解': _format_score_breakdown(c),
                '评分解释': _format_score_explanation(c),
                '命中证据': _format_keyword_evidence(c),
                # ④ 简历二次评估
                '简历评估': (
                    f"+{c['resume_eval_adjustment']}" if c.get('resume_eval_adjustment') and c['resume_eval_adjustment'] > 0
                    else str(c.get('resume_eval_adjustment', '')) if c.get('resume_eval_adjustment') is not None
                    else ''
                ),
                '简历评估理由': c.get('resume_eval_reason', ''),
                # ⑤ 跟进
                '是否打招呼': '是' if c.get('greet_sent', False) else '否',
                '跟进状态': c.get('followup_status') or ('已打招呼' if c.get('greet_sent', False) else '未沟通'),
                '跟进备注': c.get('followup_note', ''),
                '跟进时间': c.get('followup_updated_at', ''),
                '人工反馈': c.get('feedback_status', ''),
                '反馈备注': c.get('feedback_note', ''),
                '反馈时间': c.get('feedback_updated_at', ''),
                '是否需人工确认': '是' if c.get('manual_review_required') else '否',
                '风险提示': _format_risk_flags(c),
                '自动打招呼阻断原因': c.get('auto_greet_blocked_reason', ''),
                # ⑤ 原始
                '批次': c.get('batch_timestamp', ''),
                '详细信息': c.get('summary', '')
            }
            data.append(row)

        # 列顺序：身份 → 画像 → 评估 → 简历二次评估 → 跟进 → 原始
        columns = [
            '序号', '岗位', '姓名', 'geek_id',
            '年龄', '工作年限', '学历', '学历明细', '薪资', '求职状态', '城市', '最近公司', '技能',
            '匹配分', '推荐指数', '技能匹配', '评分拆解', '评分解释', '命中证据',
            '简历评估', '简历评估理由',
            '是否打招呼', '跟进状态', '跟进备注', '跟进时间',
            '人工反馈', '反馈备注', '反馈时间',
            '是否需人工确认', '风险提示', '自动打招呼阻断原因',
            '批次', '详细信息',
        ]
        # XML 1.0 合法字符集：#x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
        # BOSS API 的工作职责/个人优势字段偶尔混入垂直制表符 (\x0b) / 换页符 (\x0c) 等控制字符，
        # 直接写 openpyxl 会触发 CellILLEGAL_CHARS 校验失败，整张 Excel 导出崩掉。
        _ILLEGAL_XML_RE = re.compile(
            '[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f'
            '\ud800-\udfff﷐-﷯￾￿]'
        )

        def _clean_cell(value):
            if isinstance(value, str):
                return _ILLEGAL_XML_RE.sub('', value)
            return value

        def _write_rows(ws, header: list[str], sheet_rows: list) -> None:
            ws.append(header)
            for row in sheet_rows:
                if isinstance(row, dict):
                    ws.append([_clean_cell(row.get(col, '')) for col in header])
                else:
                    ws.append([_clean_cell(cell) for cell in row])

        def _safe_sheet_name(job_name: str) -> str:
            return str(job_name).translate(str.maketrans({
                '\\': '-', '/': '-', '*': '-', '?': '-',
                '[': '(', ']': ')', ':': '-'
            }))[:31] or "未命名岗位"

        wb = Workbook()
        ws_all = wb.active
        ws_all.title = '全部候选人'
        _write_rows(ws_all, columns, data)

        rows_by_job: dict[str, list[dict[str, Any]]] = {}
        for row in data:
            rows_by_job.setdefault(row['岗位'], []).append(row)

        for job_name, job_rows in rows_by_job.items():
            value_rows = []
            for idx, row in enumerate(job_rows, start=1):
                value_rows.append([idx] + [row[col] for col in columns[1:]])
            _write_rows(wb.create_sheet(_safe_sheet_name(job_name)), columns, value_rows)

        summary_data = []
        for job_name, job_rows in rows_by_job.items():
            rec_counts: dict[str, int] = {}
            greet_count = 0
            followup_counts: dict[str, int] = {}
            feedback_counts: dict[str, int] = {}
            scores: list[float] = []
            for row in job_rows:
                rec = row.get('推荐指数', '')
                rec_counts[rec] = rec_counts.get(rec, 0) + 1
                if row.get('是否打招呼') == '是':
                    greet_count += 1
                fu = row.get('跟进状态', '')
                if fu:
                    followup_counts[fu] = followup_counts.get(fu, 0) + 1
                fb = row.get('人工反馈', '')
                if fb:
                    feedback_counts[fb] = feedback_counts.get(fb, 0) + 1
                score = row.get('匹配分', 0)
                if isinstance(score, (int, float)):
                    scores.append(score)
            avg_score = sum(scores) / len(scores) if scores else 0
            summary_data.append({
                '岗位': job_name,
                '总人数': len(job_rows),
                '强烈推荐': rec_counts.get('强烈推荐', 0),
                '推荐': rec_counts.get('推荐', 0),
                '待定': rec_counts.get('待定', 0),
                '已打招呼': greet_count,
                '已回复': followup_counts.get('已回复', 0),
                '待约面': followup_counts.get('待约面', 0),
                '已约面': followup_counts.get('已约面', 0),
                '不合适': followup_counts.get('不合适', 0),
                '反馈合适': feedback_counts.get('合适', 0),
                '反馈误推': feedback_counts.get('误推', 0),
                '反馈放弃': feedback_counts.get('放弃', 0),
                '平均分': f"{avg_score:.1f}",
            })

        if summary_data:
            summary_columns = [
                '岗位', '总人数', '强烈推荐', '推荐', '待定', '已打招呼',
                '已回复', '待约面', '已约面', '不合适', '反馈合适', '反馈误推',
                '反馈放弃', '平均分',
            ]
            _write_rows(wb.create_sheet('统计摘要'), summary_columns, summary_data)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 设置列宽（按新列序：身份→画像→评估→跟进→原始）
            column_widths = {
                'A': 6,   # 序号
                'B': 20,  # 岗位
                'C': 10,  # 姓名
                'D': 16,  # geek_id
                'E': 8,   # 年龄
                'F': 10,  # 工作年限
                'G': 8,   # 学历
                'H': 30,  # 学历明细
                'I': 12,  # 薪资
                'J': 10,  # 求职状态
                'K': 10,  # 城市
                'L': 20,  # 最近公司
                'M': 30,  # 技能
                'N': 10,  # 匹配分
                'O': 12,  # 推荐指数
                'P': 12,  # 技能匹配
                'Q': 28,  # 评分拆解
                'R': 55,  # 评分解释
                'S': 55,  # 命中证据
                'T': 10,  # 简历评估
                'U': 55,  # 简历评估理由
                'V': 12,  # 是否打招呼
                'W': 12,  # 跟进状态
                'X': 30,  # 跟进备注
                'Y': 16,  # 跟进时间
                'Z': 10,  # 人工反馈
                'AA': 30, # 反馈备注
                'AB': 16, # 反馈时间
                'AC': 16, # 是否需人工确认
                'AD': 30, # 风险提示
                'AE': 28, # 自动打招呼阻断原因
                'AF': 16, # 批次
                'AG': 80, # 详细信息
            }
            for col, width in column_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width

            # 冻结首行（标题行）
            ws.freeze_panes = 'A2'

            # 启用自动筛选
            ws.auto_filter.ref = ws.dimensions

            # 动态查找列索引（按表头名称，不依赖固定列序）
            recommend_col = None
            greet_col = None
            for cell in ws[1]:
                if cell.value == '推荐指数':
                    recommend_col = cell.column
                elif cell.value == '是否打招呼':
                    greet_col = cell.column

            if OPENPYXL_AVAILABLE:
                if recommend_col:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=recommend_col)
                        if cell.value == "强烈推荐":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif cell.value == "推荐":
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        elif cell.value == "待定":
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

                if greet_col:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=greet_col)
                        if cell.value == "是":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        wb.save(filename)
        return True
    except Exception as e:
        print(f"Excel 导出失败：{e}")
        return False


def get_iframe(page: ChromiumPage):
    """获取包含推荐列表的 iframe"""
    try:
        frames = page.eles(_sel('iframe', 'selector', 'tag:iframe'))
        for frame in frames:
            src = frame.attr('src') or ''
            if _sel('iframe', 'src_match', 'recommend') in src.lower():
                return frame
        if frames:
            return frames[0]
    except Exception as e:
        logger.error("获取 iframe 失败：%s", e)
    return None


def extract_name_from_card(card_element: Any) -> str:
    """从候选人卡片中提取姓名"""
    try:
        # 方法 1: 查找 class=name 的独立元素
        name_elements = card_element.eles(_sel('name_extraction', 'name_xpath',
            'xpath:.//*[contains(@class, "name") and not(contains(@class, "wrap"))]'))
        if name_elements:
            for ne in name_elements:
                text = ne.text.strip() if ne.text else ""
                # 验证是否是有效的中文姓名（2-4 个汉字）
                if text and len(text) >= 2 and len(text) <= 4:
                    # 检查是否都是汉字
                    if all('一' <= c <= '鿿' for c in text):
                        return text

        # 方法 2: 从 col-2 元素的第一行提取（姓名通常在开头）
        col2_elements = card_element.eles(_sel('name_extraction', 'col2_xpath',
            'xpath:.//*[contains(@class, "col-2")]'))
        if col2_elements:
            col2_text = col2_elements[0].text.strip() if col2_elements[0].text else ""
            if col2_text:
                # 第一行可能包含姓名和职位
                first_line = col2_text.split('\n')[0].strip()
                # 提取开头 2-4 个汉字
                if len(first_line) >= 2:
                    potential_name = ""
                    for c in first_line:
                        if '一' <= c <= '鿿':
                            potential_name += c
                            if len(potential_name) > 4:
                                break
                        else:
                            break

                    if 2 <= len(potential_name) <= 4:
                        invalid_names = ['推荐', '位置', '面议', '优势', '推荐牛人', '编组',
                                       '备份', '立即', '沟通', '聊天', '联系', '在线',
                                       '今日', '最近', '邀请', '投递', '刷新', '收藏',
                                       '点赞', '分享', '高级', '资深', '初级', '中级',
                                       '离职', '在职', '看看', '沟通', '职位']
                        if potential_name not in invalid_names:
                            return potential_name

    except Exception as e:
        pass

    return "未知"


def scroll_in_frame(frame: Any, scroll_amount: int = SCROLL_PX) -> None:
    """在 iframe 内部滚动"""
    try:
        frame.run_js(f'window.scrollBy(0, {scroll_amount})')
        return True
    except Exception as e:
        logger.error("iframe 滚动失败：%s", e)
        return False


def get_frame_scroll_info(frame: Any) -> dict[str, Any]:
    """获取 iframe 内部的滚动信息"""
    try:
        scroll_top = frame.run_js('return document.documentElement.scrollTop || document.body.scrollTop || 0')
        scroll_height = frame.run_js('return document.documentElement.scrollHeight || document.body.scrollHeight || 0')
        client_height = frame.run_js('return window.innerHeight || document.documentElement.clientHeight || 0')
        return {
            'scrollTop': scroll_top,
            'scrollHeight': scroll_height,
            'clientHeight': client_height,
            'atBottom': scroll_top + client_height >= scroll_height - 50
        }
    except Exception as e:
        logger.error("获取滚动信息失败：%s", e)
        return None


def _extract_cards_batch(target):
    """单次 JS 调用批量提取所有候选人卡片数据（替代逐卡片的 N+1 调用）

    在 JS 端完成：遍历所有 [data-geekid] 卡片，提取 geek_id、innerText、姓名。
    姓名提取逻辑复用 extract_name_from_card() 的两种策略。

    Args:
        target: DrissionPage 的 page 或 iframe 对象

    Returns:
        list[dict]: [{'geek_id': str, 'name': str, 'text': str}, ...]
    """
    script = '''
    return (function(){
        var cards = document.querySelectorAll('[data-geekid]');
        var results = [];
        var invalidNames = '推荐,位置,面议,优势,推荐牛人,编组,备份,立即,沟通,聊天,联系,在线,今日,最近,邀请,投递,刷新,收藏,点赞,分享,高级,资深,初级,中级,离职,在职,看看,职位'.split(',');
        for (var c = 0; c < cards.length; c++) {
            var card = cards[c];
            var geekId = card.getAttribute('data-geekid');
            if (!geekId) continue;
            var rawText = card.innerText || '';
            if (rawText.length < 30) continue;
            var name = '';
            var nameEls = card.querySelectorAll('[class*="name"]:not([class*="wrap"])');
            for (var i = 0; i < nameEls.length; i++) {
                var t = (nameEls[i].innerText || '').trim();
                if (t.length >= 2 && t.length <= 4 && /^[\\u4e00-\\u9fff]+$/.test(t)) {
                    name = t; break;
                }
            }
            if (!name) {
                var col2 = card.querySelectorAll('[class*="col-2"]');
                if (col2.length > 0) {
                    var ct = (col2[0].innerText || '').trim();
                    var fl = ct.split('\\n')[0].trim();
                    var pn = '';
                    for (var j = 0; j < fl.length; j++) {
                        var code = fl.charCodeAt(j);
                        if (code >= 0x4e00 && code <= 0x9fff) {
                            pn += fl[j];
                            if (pn.length > 4) break;
                        } else break;
                    }
                    if (pn.length >= 2 && pn.length <= 4 && invalidNames.indexOf(pn) === -1) {
                        name = pn;
                    }
                }
            }
            results.push({geek_id: geekId, name: name || '未知', text: rawText});
        }
        return results;
    })()
    '''
    try:
        return target.run_js(script) or []
    except Exception as e:
        print(f"批量提取候选人数据失败：{e}")
        return []


def _stringify_api_value(value: Any) -> str:
    """把 API 字段安全转换为可参与文本匹配的字符串。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        parts = [_stringify_api_value(v) for v in value]
        return "、".join(p for p in parts if p)
    if isinstance(value, dict):
        for key in ("name", "label", "text", "content", "value", "desc", "title"):
            text = _stringify_api_value(value.get(key))
            if text:
                return text
        parts = [_stringify_api_value(v) for v in value.values()]
        return "、".join(p for p in parts if p)
    return str(value).strip()


def _pick_api_text(data: dict[str, Any], *keys: str) -> str:
    """按候选字段名取第一个非空文本。"""
    for key in keys:
        text = _stringify_api_value(data.get(key))
        if text:
            return text
    return ""


def _extract_geek_id_from_card(geek_card: dict[str, Any]) -> str:
    """从 BOSS 推荐接口候选人对象中提取候选人 ID。"""
    for key in ("encryptGeekId", "encGeekId", "encryptedGeekId", "geek_id", "geekId", "uid", "id"):
        value = geek_card.get(key)
        if value is not None:
            text = str(value).strip()
            if text:
                return text
    return ""


def _build_candidate_summary_from_geek_card(geek_card: dict[str, Any]) -> str:
    """把 API 返回的 geekCard 拼成现有筛选器可理解的候选人摘要文本。"""
    lines: list[str] = []

    def add(label: str, value: Any) -> None:
        text = _stringify_api_value(value)
        if text:
            lines.append(f"{label}：{text}")

    salary_text = _pick_api_text(geek_card, "salary", "expectSalaryName", "expectSalaryDesc", "expectSalary")
    if not salary_text:
        salary_text = "面议"  # API 未返回薪资时默认面议（BOSS 直聘常见）
    add("期望薪资", salary_text)
    add("姓名", _pick_api_text(geek_card, "geekName", "name", "encryptGeekName"))
    add("性别", _pick_api_text(geek_card, "geekGender", "gender", "genderDesc"))
    add("年龄", _pick_api_text(geek_card, "ageDesc", "age"))
    add("学历", _pick_api_text(geek_card, "geekDegree", "degreeName", "degree"))
    add("经验", _pick_api_text(geek_card, "geekWorkYear", "workYear", "workYearDesc"))
    add("期望职位", _pick_api_text(geek_card, "expectPositionName", "expectPosition"))
    add("期望城市", _pick_api_text(geek_card, "expectLocationName", "expectLocation"))
    add("求职状态", _pick_api_text(geek_card, "applyStatusDesc", "jobStatus", "jobStatusDesc"))

    geek_desc = geek_card.get("geekDesc")
    if isinstance(geek_desc, dict):
        add("个人优势", _pick_api_text(geek_desc, "content", "desc", "text"))
    else:
        add("个人优势", geek_desc)

    edus = geek_card.get("geekEdus") or geek_card.get("geekEduList") or []
    if isinstance(edus, list):
        for edu in edus:
            if not isinstance(edu, dict):
                add("教育经历", edu)
                continue
            school = _pick_api_text(edu, "school", "schoolName")
            major = _pick_api_text(edu, "major", "majorName")
            degree = _pick_api_text(edu, "degreeName", "degree")
            start = _pick_api_text(edu, "startDate", "startTime")
            end = _pick_api_text(edu, "endDate", "endTime")
            add("教育经历", " ".join(p for p in (school, major, degree, start, end) if p))

    works = geek_card.get("geekWorks") or geek_card.get("geekWorkList") or []
    if isinstance(works, list):
        for work in works:
            if not isinstance(work, dict):
                add("工作经历", work)
                continue
            company = _pick_api_text(work, "company", "companyName")
            position = _pick_api_text(work, "positionName", "position", "title")
            category = _pick_api_text(work, "positionCategory", "positionCategoryName")
            start = _pick_api_text(work, "startDate", "startTime")
            end = _pick_api_text(work, "endDate", "endTime")
            responsibility = _pick_api_text(work, "responsibility", "workContent", "content", "description")
            emphasis = _stringify_api_value(work.get("workEmphasisList") or work.get("workEmphasis"))
            add("工作经历", " ".join(p for p in (company, position, category, start, end) if p))
            add("工作职责", responsibility)
            add("技能标签", emphasis)

    return "\n".join(lines)


def _build_api_profile(geek_card: dict[str, Any]) -> dict[str, Any]:
    """从 geekCard 提取结构化画像，供 LLM 评估直接使用，避免从文本二次解析。

    返回字段：
        educations: [{school, major, degree, start, end}]
        works: [{company, position, category, start, end, responsibility, skills}]
        personal_summary: str
    """
    profile: dict[str, Any] = {}

    # 教育经历
    edus = geek_card.get("geekEdus") or geek_card.get("geekEduList") or []
    edu_list: list[dict[str, str]] = []
    if isinstance(edus, list):
        for edu in edus:
            if not isinstance(edu, dict):
                continue
            item = {
                k: _stringify_api_value(edu.get(field))
                for k, field in [
                    ("school", "schoolName"), ("major", "majorName"),
                    ("degree", "degreeName"), ("start", "startDate"), ("end", "endDate"),
                ]
            }
            # fallback 短键名
            for key in ("school", "major", "degree", "start", "end"):
                if not item[key]:
                    item[key] = _stringify_api_value(edu.get(key))
            # 过滤全空条目
            if any(item.values()):
                edu_list.append(item)
    profile["educations"] = edu_list

    # 工作经历
    works_raw = geek_card.get("geekWorks") or geek_card.get("geekWorkList") or []
    work_list: list[dict[str, Any]] = []
    if isinstance(works_raw, list):
        for work in works_raw:
            if not isinstance(work, dict):
                continue
            item: dict[str, Any] = {
                k: _stringify_api_value(work.get(field))
                for k, field in [
                    ("company", "companyName"), ("position", "positionName"),
                    ("category", "positionCategoryName"), ("start", "startDate"),
                    ("end", "endDate"), ("responsibility", "responsibility"),
                ]
            }
            # fallback 短键名
            for key in ("company", "position", "category", "start", "end", "responsibility"):
                if not item[key]:
                    item[key] = _stringify_api_value(work.get(key))
            # 技能标签
            skills_src = work.get("workEmphasisList") or work.get("workEmphasis") or []
            skills: list[str] = []
            if isinstance(skills_src, list):
                for s in skills_src:
                    name = s.get("name", "") if isinstance(s, dict) else str(s)
                    name = name.strip()
                    if name:
                        skills.append(name)
            item["skills"] = skills
            # 过滤全空条目
            if any(item.values()) or skills:
                work_list.append(item)
    profile["works"] = work_list

    # 个人优势
    geek_desc = geek_card.get("geekDesc")
    if isinstance(geek_desc, dict):
        profile["personal_summary"] = _stringify_api_value(
            _pick_api_text(geek_desc, "content", "desc", "text")
        )
    else:
        profile["personal_summary"] = _stringify_api_value(geek_desc)

    return profile


def _find_geek_cards_in_payload(payload: Any) -> list[dict[str, Any]]:
    """从接口响应中递归寻找 geekCard 对象。"""
    cards: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        geek_card = payload.get("geekCard")
        if isinstance(geek_card, dict):
            merged = {**geek_card}
            for key in ("encryptGeekId", "encGeekId", "encryptedGeekId"):
                if payload.get(key) and not merged.get(key):
                    merged[key] = payload[key]
            cards.append(merged)
        elif _extract_geek_id_from_card(payload):
            cards.append(payload)

        for list_key in ("list", "geekList"):
            data_list = payload.get(list_key)
            if isinstance(data_list, list):
                for item in data_list:
                    cards.extend(_find_geek_cards_in_payload(item))

        for data_key in ("data", "zpData"):
            data = payload.get(data_key)
            if isinstance(data, (dict, list)):
                cards.extend(_find_geek_cards_in_payload(data))

        result = payload.get("result")
        if isinstance(result, (dict, list)):
            cards.extend(_find_geek_cards_in_payload(result))

        friend_list = payload.get("friendList")
        if isinstance(friend_list, list):
            for item in friend_list:
                cards.extend(_find_geek_cards_in_payload(item))
    elif isinstance(payload, list):
        for item in payload:
            cards.extend(_find_geek_cards_in_payload(item))
    return cards


def _extract_candidates_from_api_payload(payload: Any) -> list[dict[str, str]]:
    """从 BOSS 推荐接口 JSON 中提取候选人，输出兼容 DOM 提取的结构。"""
    candidates = []
    for geek_card in _find_geek_cards_in_payload(payload):
        geek_id = _extract_geek_id_from_card(geek_card)
        if not geek_id:
            continue
        summary = _build_candidate_summary_from_geek_card(geek_card)
        if not summary:
            continue
        # 提取结构化字段，避免下游用正则从文本中重新解析
        structured: dict[str, Any] = {}
        work_year_text = _pick_api_text(geek_card, "geekWorkYear", "workYear", "workYearDesc")
        if work_year_text:
            parsed = parse_experience_years(work_year_text)
            if parsed is not None:
                structured['exp_years'] = parsed
        age_text = _pick_api_text(geek_card, "ageDesc", "age")
        if age_text:
            m = re.search(r'(\d+)', age_text)
            if m:
                structured['age'] = int(m.group(1))
        degree_text = _pick_api_text(geek_card, "geekDegree", "degreeName", "degree")
        if degree_text:
            structured['degree'] = degree_text
        city_text = _pick_api_text(geek_card, "expectLocationName", "expectLocation")
        if city_text:
            structured['city'] = city_text
        job_status_text = _pick_api_text(geek_card, "applyStatusDesc", "jobStatus", "jobStatusDesc")
        if job_status_text:
            structured['job_status'] = job_status_text
        salary_text = _pick_api_text(geek_card, "salary", "expectSalaryName", "expectSalaryDesc", "expectSalary")
        if salary_text and '面议' not in salary_text:
            # 范围：15-25K / 15K-25K / 15-25（K 可选）
            m = re.search(r'(\d+)\s*[kK]?\s*[-~～\-]\s*(\d+)\s*[kK]?', salary_text)
            if m:
                structured['salary_min'] = int(m.group(1))
                structured['salary_max'] = int(m.group(2))
            else:
                # 单值：15K / 15薪 / 15千
                m = re.search(r'(\d+)\s*[kK薪千]', salary_text)
                if m:
                    structured['salary_min'] = int(m.group(1))
                    structured['salary_max'] = int(m.group(1))
        candidates.append({
            "geek_id": geek_id,
            "name": _pick_api_text(geek_card, "geekName", "name", "encryptGeekName") or "未知",
            "summary": summary,
            "structured": structured,
            "_api_profile": _build_api_profile(geek_card),
        })
    return candidates


def _merge_candidates_into_list(
    batch: list[dict[str, Any]],
    all_candidates: list[dict[str, Any]],
    seen_geek_ids: set[str],
    candidate_index_by_id: dict[str, int],
) -> list[dict[str, Any]]:
    """将一批候选人合并到累计列表，去重，返回本批新增的候选人。"""
    candidates_in_round: list[dict[str, Any]] = []
    current_round_ids: set[str] = set()
    for item in batch:
        geek_id = item.get('geek_id', '') or ''
        if not geek_id:
            continue
        if geek_id in seen_geek_ids or geek_id in current_round_ids:
            existing_idx = candidate_index_by_id.get(geek_id)
            if item.get('_source') == 'api' and existing_idx is not None:
                existing = all_candidates[existing_idx]
                new_summary = item.get('summary', '')
                if new_summary and len(new_summary) > len(existing.get('summary', '')):
                    existing['summary'] = new_summary
                    existing['name'] = item.get('name') or existing.get('name', '未知')
                if item.get('structured') and not existing.get('structured'):
                    existing['structured'] = item['structured']
                if item.get('_api_profile') and not existing.get('_api_profile'):
                    existing['_api_profile'] = item['_api_profile']
            continue

        current_round_ids.add(geek_id)
        text = item.get('summary', '') or ''
        # API 来源已有结构化数据，跳过文本内容检查
        is_api_with_structured = (item.get('_source') == 'api' and item.get('structured'))
        if not is_api_with_structured:
            has_candidate_info = (
                '经验' in text or '本科' in text or '硕士' in text or
                'Java' in text or '开发' in text or '工程师' in text or
                re.search(r'\d+年', text) or re.search(r'\d+岁', text)
            )
            if not has_candidate_info:
                continue

        candidate = {
            'geek_id': geek_id,
            'name': item.get('name', '未知'),
            'summary': text,
            'structured': item.get('structured'),
        }
        if item.get('_api_profile'):
            candidate['_api_profile'] = item['_api_profile']
        candidates_in_round.append(candidate)

    for c in candidates_in_round:
        candidate_index_by_id[c['geek_id']] = len(all_candidates)
        all_candidates.append(c)
        seen_geek_ids.add(c['geek_id'])

    return candidates_in_round


def _merge_api_enrichment_into_existing(
    api_candidates: list[dict[str, Any]],
    all_candidates: list[dict[str, Any]],
    candidate_index_by_id: dict[str, int],
    source_tag: str = '',
) -> int:
    """Use API/listener data only to enrich candidates already discovered in DOM.

    Args:
        source_tag: 可选来源标记（'listener' / 'api_fallback'），首次成功合并
            structured 数据时写入候选人的 ``_enriched_by`` 字段，用于最终统计
            listener 和 API 兜底各自的贡献比例。已存在标记的候选人不会被覆盖
            （保持"首次写入者胜"语义）。
    """
    updated = 0
    for item in api_candidates:
        geek_id = item.get('geek_id', '') or ''
        if not geek_id:
            continue
        existing_idx = candidate_index_by_id.get(geek_id)
        if existing_idx is None:
            continue

        existing = all_candidates[existing_idx]
        changed = False

        new_summary = item.get('summary', '') or ''
        if new_summary and len(new_summary) > len(existing.get('summary', '') or ''):
            existing['summary'] = new_summary
            existing['name'] = item.get('name') or existing.get('name', '未知')
            changed = True

        new_structured = item.get('structured') or {}
        if new_structured:
            current_structured = existing.get('structured') or {}
            merged_structured = dict(current_structured)
            before_structured = dict(merged_structured)
            for key, value in new_structured.items():
                if value not in (None, '') and not merged_structured.get(key):
                    merged_structured[key] = value
            if merged_structured != before_structured:
                existing['structured'] = merged_structured
                if source_tag and not existing.get('_enriched_by'):
                    existing['_enriched_by'] = source_tag
                changed = True

        if item.get('_api_profile') and not existing.get('_api_profile'):
            existing['_api_profile'] = item['_api_profile']
            changed = True

        if changed:
            updated += 1

    return updated

class _ApiCapture:
    """基于 JS fetch 拦截的 API 响应捕获器。

    通过注入 JS 拦截 fetch() 调用，捕获推荐接口的请求和响应。
    比 CDP 回调更可靠，能拿到完整的 URL（含 query string）和 POST body。
    """

    # 注入到页面的 JS 代码：拦截 fetch + XMLHttpRequest，缓存匹配的响应
    _INJECT_JS = '''
    if (!window.__bossApiCapture) {
        window.__bossApiCapture = {requests: [], origFetch: window.fetch, injected: true};

        // 拦截 fetch
        window.fetch = async function(...args) {
            const resp = await window.__bossApiCapture.origFetch.apply(this, args);
            try {
                const url = (typeof args[0] === 'string') ? args[0] : (args[0]?.url || '');
                if (url.includes('geek') && (url.includes('/wapi/') || url.includes('/api/'))) {
                    const clone = resp.clone();
                    const body = await clone.text();
                    window.__bossApiCapture.requests.push({url: url, body: body, method: args[1]?.method || 'GET'});
                }
            } catch(e) {}
            return resp;
        };

        // 拦截 XMLHttpRequest
        var origOpen = XMLHttpRequest.prototype.open;
        var origSend = XMLHttpRequest.prototype.send;
        XMLHttpRequest.prototype.open = function(method, url) {
            this.__bossUrl = url || '';
            this.__bossMethod = method;
            return origOpen.apply(this, arguments);
        };
        XMLHttpRequest.prototype.send = function(body) {
            var xhr = this;
            var url = xhr.__bossUrl || '';
            if (url.includes('geek') && (url.includes('/wapi/') || url.includes('/api/'))) {
                xhr.addEventListener('load', function() {
                    try {
                        window.__bossApiCapture.requests.push({
                            url: url, body: xhr.responseText, method: xhr.__bossMethod || 'GET'
                        });
                    } catch(e) {}
                });
            }
            return origSend.apply(this, arguments);
        };
    }
    '''

    def __init__(self):
        self._page = None
        self._iframe = None
        self._active = False

    def start(self, page: ChromiumPage) -> bool:
        """注入 JS fetch 拦截器（优先注入 iframe，因为 API 请求在 iframe 内发出）。"""
        try:
            iframe = get_iframe(page)
            target = iframe if iframe else page
            target.run_js(self._INJECT_JS)
            # 验证注入是否生效
            injected = target.run_js('return !!window.__bossApiCapture && window.__bossApiCapture.injected')
            if not injected:
                page.run_js(self._INJECT_JS)
                target = page
                injected = page.run_js('return !!window.__bossApiCapture && window.__bossApiCapture.injected')
            self._page = page
            self._iframe = target if target is not page else None
            self._active = bool(injected)
            return self._active
        except Exception:
            pass
        return False

    def consume(self) -> tuple[list[dict[str, str]], str]:
        """读取并清空 JS 缓存的 API 响应，返回 (candidates, api_url)。"""
        if not self._active:
            return [], ""
        target = self._iframe if self._iframe else self._page
        if not target:
            return [], ""
        try:
            raw = target.run_js('''
                (function() {
                    const c = window.__bossApiCapture;
                    if (!c) return '[]';
                    const reqs = c.requests.splice(0, c.requests.length);
                    return JSON.stringify(reqs);
                })()
            ''')
            if not raw:
                return [], ""
            import json as _json
            items = _json.loads(raw)
        except Exception:
            return [], ""

        candidates: list[dict[str, str]] = []
        api_url = ""
        for item in items:
            if not api_url:
                api_url = item.get("url", "")
            body = item.get("body", "")
            if body:
                try:
                    payload = _json.loads(body)
                    candidates.extend(_extract_candidates_from_api_payload(payload))
                except Exception:
                    pass
        return candidates, api_url

    def stop(self):
        """停止拦截（刷新页面后自动失效，无需显式还原）。"""
        self._active = False


def _start_recommend_api_listener(page: ChromiumPage) -> Any | None:
    """启动推荐列表接口监听，优先使用 DrissionPage 原生网络监听。"""
    try:
        listener = page.listen
        try:
            listener.stop()
        except Exception:
            pass
        # BOSS 会调整推荐接口路径；宽监听 geek 相关 XHR，再由 payload parser 判定是否有候选人。
        listener.start("geek", method=("GET", "POST"), res_type=("XHR", "Fetch"))
        return listener
    except Exception:
        pass

    capture = _ApiCapture()
    if capture.start(page):
        return capture
    print("启动 API 监听失败，将使用页面提取")
    return None


def _consume_recommend_api_candidates(listener: Any | None, timeout: float = 0.05) -> tuple[list[dict[str, str]], str]:
    """消费已监听到的推荐接口响应，提取候选人数据。"""
    if not listener:
        return [], ""

    # _ApiCapture 类型
    if isinstance(listener, _ApiCapture):
        return listener.consume()

    # DrissionPage Listener 类型
    candidates: list[dict[str, str]] = []
    api_url = ""
    while True:
        try:
            packet = listener.wait(timeout=timeout, fit_count=False)
        except Exception as e:
            logger.error("读取 API 监听数据失败：%s", e)
            break
        if not packet:
            break

        packets = packet if isinstance(packet, list) else [packet]
        for p in packets:
            try:
                if getattr(p, "is_failed", False):
                    continue
                if not api_url:
                    api_url = getattr(p, "url", "") or ""
                payload = p.response.body
                candidates.extend(_extract_candidates_from_api_payload(payload))
            except Exception as e:
                logger.error("解析 API 监听数据失败：%s", e)
                continue

        timeout = 0.01

    return candidates, api_url


def _looks_like_recommend_api_url(url: str) -> bool:
    """Best-effort match for BOSS recommendation API URLs."""
    lowered = (url or "").lower()
    return (
        "geek" in lowered
        and ("/wapi/" in lowered or "/api/" in lowered)
        and any(token in lowered for token in ("rec", "recommend", "list"))
    )


def _find_recent_recommend_api_url(*targets: Any) -> str:
    """Find a recent geek recommendation XHR/fetch URL from browser performance entries."""
    seen: set[str] = set()
    urls: list[str] = []
    js = r'''
        return (function() {
            try {
                return JSON.stringify(
                    performance.getEntriesByType('resource')
                        .map(e => e && e.name || '')
                        .filter(Boolean)
                        .filter(u => u.indexOf('geek') >= 0 && (u.indexOf('/wapi/') >= 0 || u.indexOf('/api/') >= 0))
                        .slice(-20)
                );
            } catch (e) {
                return '[]';
            }
        })()
    '''
    for target in targets:
        if not target:
            continue
        try:
            raw = target.run_js(js)
            import json as _json
            for url in _json.loads(raw or "[]"):
                if url and url not in seen:
                    urls.append(url)
                    seen.add(url)
        except Exception:
            continue

    preferred = [url for url in urls if _looks_like_recommend_api_url(url)]
    if preferred:
        return preferred[-1]
    return ""


def _parse_api_pagination(url: str) -> dict[str, Any] | None:
    """从 BOSS 推荐接口 URL 中提取分页参数，用于后续直调。

    BOSS 常见分页格式：
      - page 参数：?page=1&pageSize=20
      - cursor 参数：?cursor=abc123

    Returns:
        {'base_url': str, 'page_param': str, 'page_size': int, 'query_params': dict} or None
    """
    if not url:
        return None
    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
    parsed = urlparse(url)
    params = parse_qs(parsed.query, keep_blank_values=True)
    # 扁平化单值参数
    flat_params = {k: v[0] for k, v in params.items()}

    page_param = None
    page_size = 20
    if 'page' in flat_params:
        page_param = 'page'
        page_size = int(flat_params.get('pageSize', flat_params.get('page_size', 20)))
    elif 'cursor' in flat_params:
        page_param = 'cursor'
        page_size = int(flat_params.get('pageSize', flat_params.get('page_size', 20)))

    if page_param is None:
        return None

    base = urlunparse((parsed.scheme, parsed.netloc, parsed.path, '', '', ''))
    return {
        'base_url': base,
        'page_param': page_param,
        'page_size': page_size,
        'query_params': flat_params,
    }


def _build_recommend_api_pagination_from_page(target: Any) -> dict[str, Any] | None:
    """从当前推荐页 iframe URL 构造推荐接口分页参数，不刷新页面。"""
    try:
        href = target.run_js('return location.href') or ""
    except Exception:
        return None

    from urllib.parse import urlparse, parse_qs, urlunparse
    parsed = urlparse(href)
    params = parse_qs(parsed.query, keep_blank_values=True)
    job_id = (params.get('jobid') or params.get('jobId') or [""])[0]
    if not job_id:
        return None

    base = urlunparse((parsed.scheme, parsed.netloc, "/wapi/zpjob/rec/geek/list", '', '', ''))
    return {
        'base_url': base,
        'page_param': 'page',
        'page_size': None,
        'query_params': {
            'age': '16,-1',
            'activation': '0',
            'school': '0',
            'gender': '0',
            'recentNotView': '0',
            'exchangeResumeWithColleague': '0',
            'major': '0',
            'keyword1': '-1',
            'switchJobFrequency': '0',
            'degree': '0',
            'experience': '0',
            'intention': '0',
            'salary': '0',
            'jobId': job_id,
            'page': '1',
            'coverScreenMemory': '0',
            'cardType': '0',
        },
    }


def _read_recommend_page_identity(target: Any) -> dict[str, str]:
    """读取当前推荐 iframe 的岗位标识，用于刷新监听兜底前后校验。"""
    try:
        href = target.run_js('return location.href') or ""
    except Exception:
        href = ""

    job_id = ""
    if href:
        try:
            from urllib.parse import urlparse, parse_qs
            parsed = urlparse(href)
            params = parse_qs(parsed.query, keep_blank_values=True)
            job_id = (params.get('jobid') or params.get('jobId') or params.get('job_id') or [""])[0]
        except Exception:
            job_id = ""

    title = ""
    try:
        title = target.run_js(r'''
            return (function() {
                const text = (document.body && document.body.innerText || '').split('\n')
                    .map(s => s.trim()).filter(Boolean);
                for (let i = 0; i < text.length; i++) {
                    if (text[i] === '最新' && text[i + 1]) return text[i + 1];
                    if (text[i] === '推荐' && text[i + 1] && text[i + 1] !== '最新') return text[i + 1];
                }
                return '';
            })()
        ''') or ""
    except Exception:
        title = ""

    return {"job_id": job_id, "job_title": title, "href": href}


def _same_recommend_page_identity(before: dict[str, str], after: dict[str, str]) -> bool:
    """判断刷新前后是否仍是同一推荐岗位。

    job_id（URL 参数）是最可靠的标识，两者都读到且相等即可确认同一岗位。
    title 作为 job_id 不可用时的 fallback，刷新前能读到但刷新后读不到 → 保守视为已变化。
    """
    before_job = before.get("job_id", "")
    after_job = after.get("job_id", "")
    if before_job and after_job:
        return before_job == after_job
    if before_job and not after_job:
        return False

    before_title = before.get("job_title", "")
    after_title = after.get("job_title", "")
    if before_title and after_title and before_title != after_title:
        return False
    if before_title and not after_title:
        return False

    return True


def _is_api_risk_status(status: Any) -> bool:
    """判断接口错误是否应视为风控熔断，而不是普通接口失败。"""
    try:
        status_int = int(status)
    except (TypeError, ValueError):
        return False
    return status_int in {403, 412, 429}


def _fetch_api_page_result(page: Any, pagination: dict[str, Any], page_num: int) -> tuple[list[dict[str, str]], bool | None]:
    """通过浏览器 fetch 直接调用 BOSS 推荐接口分页，返回候选人和 hasMore。"""
    from urllib.parse import urlencode
    params = dict(pagination['query_params'])
    if pagination['page_param'] == 'page':
        params['page'] = str(page_num)
        if pagination.get('page_size'):
            params['pageSize'] = str(pagination['page_size'])
    else:
        return [], None

    query_string = urlencode(params)
    full_url = f"{pagination['base_url']}?{query_string}"

    js_code = f'''
    return (async () => {{
        try {{
            const resp = await fetch("{full_url}", {{credentials: "include"}});
            if (!resp.ok) return JSON.stringify({{error: resp.status}});
            return await resp.text();
        }} catch(e) {{
            return JSON.stringify({{error: e.message}});
        }}
    }})()
    '''

    try:
        result = page.run_js(js_code)
        if not result:
            return [], None
        import json as _json
        payload = _json.loads(result)
        if isinstance(payload, dict) and 'error' in payload:
            if _is_api_risk_status(payload['error']):
                raise ApiRiskBlocked(payload['error'], page_num)
            logger.error("  API 分页直调失败 (page=%d): %s", page_num, payload['error'])
            return [], None
        zp_data = payload.get('zpData') if isinstance(payload, dict) else {}
        has_more = zp_data.get('hasMore') if isinstance(zp_data, dict) else None
        return _extract_candidates_from_api_payload(payload), has_more
    except ApiRiskBlocked:
        raise
    except Exception as e:
        logger.error("  API 分页直调异常 (page=%d): %s", page_num, e)
        return [], None


def _fetch_api_page(page: Any, pagination: dict[str, Any], page_num: int) -> list[dict[str, str]]:
    """通过浏览器 fetch 直接调用 BOSS 推荐接口分页。

    Args:
        page: DrissionPage ChromiumPage 对象
        pagination: _parse_api_pagination 返回的分页信息
        page_num: 要请求的页码（从 1 开始）

    Returns:
        候选人列表，失败时返回空列表
    """
    candidates, _ = _fetch_api_page_result(page, pagination, page_num)
    return candidates


def extract_candidates_by_comprehensive_analysis(page, max_rounds=MAX_ROUNDS_DEFAULT, progress_callback=None, stop_event=None, captcha_callback=None, notice_callback=None, blocking_notice_callback=None, max_candidates=API_CANDIDATE_LIMIT_DEFAULT, use_api_extraction=True, extraction_mode=None):
    """通过全面分析提取候选人

    Args:
        page: 页面对象
        max_rounds: 最大滚动轮次（默认 30）
        progress_callback: 进度回调 callable(percentage, description)，percentage 0-100
        stop_event: threading.Event，设位时立即停止扫描
        captcha_callback: 验证码回调 callable(stage, detail)
        notice_callback: GUI 提示回调 callable(title, message)
        max_candidates: API 直调补全预算，用于推导最多补全页数；0 表示使用保守默认值
        use_api_extraction: 兼容旧调用；False 等价于 extraction_mode="dom"
        extraction_mode: 提取模式，api=listener+refresh+DOM+API补全，listener=listener+refresh+DOM补全，dom=仅DOM滚动
    """
    print("正在提取候选人...")
    time.sleep(_human_delay(1.0, 0.5))

    iframe = get_iframe(page)

    all_candidates = []
    seen_geek_ids = set()
    candidate_index_by_id: dict[str, int] = {}
    target = iframe if iframe else page
    consecutive_empty = 0

    if extraction_mode is None:
        extraction_mode = "api" if use_api_extraction else "dom"
    api_enrichment_enabled = extraction_mode == "api"
    listener_enabled = extraction_mode in {"api", "listener"}

    # DOM 是候选人来源；listener/API 只能按 geek_id 补全已出现的候选人。
    api_listener = _start_recommend_api_listener(page) if listener_enabled else None
    if listener_enabled and not api_listener:
        print("启动 API 监听失败，将继续使用 DOM 提取")
    pending_listener_candidates: list[dict[str, Any]] = []
    pending_listener_url = ""
    observed_api_url = ""
    listener_refresh_captured = False
    last_round_new_count = 0

    # listener 启动后刷新一次，让首屏 API 请求被监听器捕获（结构化字段来源）。
    # 刷新会使页面恢复默认岗位，用 identity 校验检测是否跑偏。
    if api_listener:
        identity_before = _read_recommend_page_identity(target)
        try:
            page.refresh()
            # 轮询等待 iframe 内容就绪（候选人卡片出现），最多 10 秒
            for _wait in range(20):
                if stop_event and stop_event.is_set():
                    break
                time.sleep(0.5)
                iframe = get_iframe(page)
                if iframe:
                    try:
                        has_cards = iframe.run_js('return !!(document.querySelector("[data-geekid]"))')
                        if has_cards:
                            break
                    except Exception:
                        pass
            else:
                iframe = get_iframe(page)
            target = iframe if iframe else page
            identity_after = _read_recommend_page_identity(target)
            # 刷新后完全读不到身份标识 → 页面未加载完，跳过校验
            if not identity_after.get("job_id") and not identity_after.get("job_title"):
                print("刷新后页面加载中，跳过岗位校验")
            elif not _same_recommend_page_identity(identity_before, identity_after):
                before_title = identity_before.get("job_title") or identity_before.get("job_id") or "原岗位"
                after_title = identity_after.get("job_title") or identity_after.get("job_id") or "当前岗位"
                msg = (f"刷新页面后岗位已变化：\n"
                       f"  原岗位：{before_title}\n"
                       f"  当前岗位：{after_title}\n\n"
                       f"请在浏览器中切回目标岗位后点击「确定」继续。")
                print(f"⚠️  刷新后岗位标识变化：{before_title} → {after_title}")
                # 先清空 listener 中错误岗位的旧数据，再弹窗；
                # 弹窗期间用户切岗产生的新 API 响应会保留在 listener 中
                _consume_recommend_api_candidates(api_listener, timeout=0.01)
                _blocker = blocking_notice_callback or notice_callback
                if _blocker:
                    _blocker("岗位已变化", msg)
                else:
                    input(f"\n⚠️  {msg}\n按 Enter 继续（或 Ctrl+C 中止）...")
            captured = _consume_recommend_api_candidates(api_listener, timeout=2.0)
            if isinstance(captured, tuple) and len(captured) == 2:
                pending_listener_candidates, pending_listener_url = captured
                listener_refresh_captured = bool(pending_listener_candidates or pending_listener_url)
                if pending_listener_url:
                    observed_api_url = pending_listener_url
            if not observed_api_url:
                observed_api_url = _find_recent_recommend_api_url(target, page)
            print(f"listener + refresh 捕获: {len(pending_listener_candidates)} 条")
        except Exception as e:
            print(f"刷新页面失败，将继续使用当前页面：{e}")

    try:
        for scroll_round in range(max_rounds):
            # 检查停止信号
            if stop_event and stop_event.is_set():
                raise StopRequested()

            if not _ensure_recommend_page(page, notice_callback=notice_callback, context="扫描候选人"):
                break

            # 验证码检测：每 3 轮一次（降低调用频率，弹窗一旦出现 1.5s 内必然可见）
            if scroll_round % 3 == 0:
                is_captcha, captcha_msg = _detect_captcha(page)
                if is_captcha:
                    print(f"\n⚠️  检测到安全验证弹窗 ({captcha_msg})")
                    if not _wait_for_captcha_resolution(page, stop_event, captcha_callback=captcha_callback, detail=captcha_msg, stage="scan"):
                        break

            # 进度上报
            if progress_callback:
                pct = int((scroll_round + 1) / max_rounds * 100)
                progress_callback(pct, f"正在扫描候选人... 第{scroll_round + 1}/{max_rounds}轮")

            # DOM 扫描始终滚动页面，确保候选人集合和后续可点击列表一致。
            if scroll_round > 0:
                if iframe:
                    # 滚动 window + 候选人列表的实际滚动容器
                    iframe.run_js(f'''
                        (function() {{
                            var winBefore = window.scrollY || 0;
                            window.scrollBy(0, {SCROLL_PX});
                            var winAfter = window.scrollY || 0;
                            // 从候选人卡片往上找真正的滚动容器
                            var card = document.querySelector('[data-geekid]');
                            var found = null;
                            if (card) {{
                                var el = card.parentElement;
                                while (el && el !== document.body && el !== document.documentElement) {{
                                    var ov = getComputedStyle(el).overflowY;
                                    if (el.scrollHeight > el.clientHeight + 10
                                        && (ov === 'auto' || ov === 'scroll')) {{
                                        var before = el.scrollTop;
                                        el.scrollTop += {SCROLL_PX};
                                        found = {{
                                            tag: el.tagName,
                                            cls: (el.className || '').substring(0, 60),
                                            before: before,
                                            after: el.scrollTop,
                                            sh: el.scrollHeight,
                                            ch: el.clientHeight,
                                            ov: ov
                                        }};
                                        break;
                                    }}
                                    el = el.parentElement;
                                }}
                            }}
                            var cards = document.querySelectorAll('[data-geekid]');
                            return {{
                                winBefore: winBefore, winAfter: winAfter,
                                container: found,
                                cardCount: cards.length
                            }};
                        }})()
                    ''')
                    time.sleep(_human_delay(0.8, 0.5))
                else:
                    page.run_js(f'window.scrollBy(0, {SCROLL_PX})')
                    time.sleep(_human_delay(0.8, 0.5))

                # 到底检测：滚动位置优先（单次 JS 调用，无 DOM 查找）
                scroll_info = get_frame_scroll_info(iframe) if iframe else None
                if scroll_info and scroll_info.get('atBottom'):
                    # 兜底：再用文本确认一次
                    bottom_hint = None
                    try:
                        _bottom_texts = _sel('scroll', 'bottom_texts', ["到底", "没有更多"])
                        bottom_hint = target.ele(f'@text():{_bottom_texts[0]}', timeout=0.3)
                        if not bottom_hint and len(_bottom_texts) > 1:
                            bottom_hint = target.ele(f'@text():{_bottom_texts[1]}', timeout=0.3)
                    except Exception:
                        pass
                    if bottom_hint:
                        print(f"检测到'到底'提示，第 {scroll_round + 1} 轮提前终止（累计 {len(all_candidates)} 个候选人）")
                        break

            # 收集候选人：DOM 建集合，listener 只增强已有 geek_id。
            candidates_in_round = []

            try:
                batch = []
                dom_batch = _extract_cards_batch(target)
                for item in dom_batch:
                    batch.append({
                        'geek_id': item.get('geek_id', ''),
                        'name': item.get('name', '未知'),
                        'summary': item.get('text', ''),
                        '_source': 'dom',
                    })

                candidates_in_round = _merge_candidates_into_list(
                    batch, all_candidates, seen_geek_ids, candidate_index_by_id
                )

                if pending_listener_candidates:
                    matched = _merge_api_enrichment_into_existing(
                        pending_listener_candidates, all_candidates, candidate_index_by_id,
                        source_tag='listener',
                    )
                    print(
                        f"listener + refresh 合并: 返回 {len(pending_listener_candidates)} 条, "
                        f"命中 DOM {matched} 条"
                    )
                    pending_listener_candidates = []
                    pending_listener_url = ""

                if api_listener and not (scroll_round == 0 and listener_refresh_captured):
                    consumed = _consume_recommend_api_candidates(api_listener)
                    if isinstance(consumed, tuple) and len(consumed) == 2:
                        api_candidates, _api_url = consumed
                        if _api_url:
                            observed_api_url = _api_url
                    else:
                        api_candidates = []
                    if api_candidates or _api_url:
                        print(
                            f"listener 滚动捕获(第 {scroll_round + 1} 轮): "
                            f"返回 {len(api_candidates)} 条"
                        )
                    matched = _merge_api_enrichment_into_existing(
                        api_candidates, all_candidates, candidate_index_by_id,
                        source_tag='listener',
                    )
                    if api_candidates or _api_url:
                        print(f"listener 滚动合并(第 {scroll_round + 1} 轮): 命中 DOM {matched} 条")

            except Exception as e:
                logger.error("提取候选人元素失败(轮次%d): %s", scroll_round + 1, e)

            new_count = len(candidates_in_round)
            last_round_new_count = new_count
            total_count = len(all_candidates)

            # 连续空轮次检测（兜底策略，不依赖特定文案）
            if new_count == 0:
                consecutive_empty += 1
                if consecutive_empty >= EMPTY_ROUNDS_LIMIT:
                    print(f"连续 {consecutive_empty} 轮无新候选人，第 {scroll_round + 1} 轮提前终止（累计 {total_count} 个候选人）")
                    break
            else:
                consecutive_empty = 0

            # 每 10 轮或最后一轮打印进度
            if (scroll_round + 1) % 10 == 0 or (scroll_round + 1) == max_rounds:
                status = f"+{new_count}" if new_count > 0 else "无新增"
                print(f"轮次 {scroll_round + 1}/{max_rounds}: {status}, 累计 {total_count} 个")
        else:
            if last_round_new_count > 0:
                warning = (
                    f"已达到扫描轮次上限 {max_rounds}，最后一轮仍新增 "
                    f"{last_round_new_count} 人，候选人可能未提取完整"
                )
                print(f"[WARN] {warning}")
    finally:
        if api_listener:
            try:
                api_listener.stop()
            except Exception:
                pass

    if api_enrichment_enabled and all_candidates:
        missing_ids = {
            c.get('geek_id') for c in all_candidates
            if c.get('geek_id') and not c.get('structured')
        }
        if missing_ids:
            if not observed_api_url:
                observed_api_url = _find_recent_recommend_api_url(target, page)
            pagination = _parse_api_pagination(observed_api_url) if observed_api_url else None
            if not pagination:
                pagination = _build_recommend_api_pagination_from_page(target)
            if pagination:
                if max_candidates and max_candidates > 0:
                    page_limit = min(20, max(5, (max_candidates + 19) // 20))
                else:
                    page_limit = 5
                print(f"API 直调仅补全 DOM 已出现候选人，最多 {page_limit} 页")
                consecutive_misses = 0
                api_limit_reached_with_hits = False
                for page_num in range(1, page_limit + 1):
                    if stop_event and stop_event.is_set():
                        raise StopRequested()
                    if page_num > 1:
                        time.sleep(_human_delay(API_PAGE_DELAY_CENTER, API_PAGE_DELAY_SPREAD))
                    try:
                        api_candidates, has_more = _fetch_api_page_result(target, pagination, page_num)
                    except ApiRiskBlocked as e:
                        print(f"API 返回疑似风控状态码 {e.status}（第 {e.page_num} 页），停止 API 补全。")
                        break
                    matched = _merge_api_enrichment_into_existing(
                        api_candidates, all_candidates, candidate_index_by_id,
                        source_tag='api_fallback',
                    )
                    print(
                        f"API 兜底第 {page_num} 页: 返回 {len(api_candidates)} 条, "
                        f"命中 DOM {matched} 条"
                    )
                    if matched == 0:
                        consecutive_misses += 1
                        if consecutive_misses >= 3:
                            print(f"API 兜底连续 {consecutive_misses} 页无新增命中，提前停止")
                            break
                    else:
                        consecutive_misses = 0
                    missing_ids = {
                        c.get('geek_id') for c in all_candidates
                        if c.get('geek_id') and not c.get('structured')
                    }
                    if not missing_ids:
                        break
                    if has_more is False:
                        break
                    if (
                        page_num == page_limit
                        and matched > 0
                        and missing_ids
                    ):
                        api_limit_reached_with_hits = True
                if api_limit_reached_with_hits:
                    print(
                        f"[WARN] API 补全已达到 {page_limit} 页上限，最后一页仍命中 "
                        f"DOM 候选人，仍有 {len(missing_ids)} 人缺少结构化信息"
                    )

    # 统计 API 结构化数据覆盖率
    _api_count = sum(1 for c in all_candidates if c.get('structured'))
    if all_candidates:
        _fallback_count = len(all_candidates) - _api_count
        print(f"结构化数据覆盖: {_api_count}/{len(all_candidates)} 人"
              f" ({_api_count * 100 // len(all_candidates)}%)")
        # 拆分 listener 与 API 兜底的贡献（首次合并者胜，已在 _merge_... 中标记）
        _listener_enriched = sum(
            1 for c in all_candidates if c.get('_enriched_by') == 'listener'
        )
        _api_fb_enriched = sum(
            1 for c in all_candidates if c.get('_enriched_by') == 'api_fallback'
        )
        coverage_details = []
        if _listener_enriched:
            coverage_details.append(("listener 结构化", _listener_enriched))
        if _api_fb_enriched:
            coverage_details.append(("API 兜底结构化", _api_fb_enriched))
        if _fallback_count:
            coverage_details.append(("文本降级解析", _fallback_count))
        for index, (label, count) in enumerate(coverage_details):
            branch = "└─" if index == len(coverage_details) - 1 else "├─"
            print(f"  {branch} {label}: {count} 人")

    print(f"\n=== 提取完成 ===")
    print(f"总共找到 {len(all_candidates)} 个候选人")
    return all_candidates


def _find_card_by_scroll(target, card_css, stop_event=None, max_scrolls=MAX_SCROLL_SEARCH, scroll_px=SCROLL_PX):
    """智能滚动定位候选人卡片

    BOSS 直聘使用虚拟列表，只渲染视口内的卡片。此函数通过系统性滚动搜索
    不在当前视口的候选人，解决补打招呼和右键打招呼找不到卡片的问题。

    策略：
    1. 先检查当前位置（最快路径，刚扫描完的场景）
    2. 滚到顶部建立已知起点
    3. 从顶部向下逐步滚动，每步检查卡片是否渲染

    Args:
        target: DrissionPage 的 page 或 iframe 对象
        card_css: CSS 选择器字符串（含 data-geekid）
        stop_event: threading.Event，设置时立即放弃搜索
        max_scrolls: 最大滚动次数（默认 40，覆盖约 32000px 的列表高度）
        scroll_px: 每次滚动像素（默认 800）

    Returns:
        card element 或 None
    """
    # Phase 1: 当前位置快速检查
    try:
        card = target.ele(card_css, timeout=1)
        if card:
            return card
    except Exception:
        pass

    # Phase 2: 滚到顶部
    try:
        target.run_js('''
            window.scrollTo(0, 0);
            var card = document.querySelector('[data-geekid]');
            if (card) {
                var el = card.parentElement;
                while (el && el !== document.body && el !== document.documentElement) {
                    var ov = getComputedStyle(el).overflowY;
                    if (el.scrollHeight > el.clientHeight + 10
                        && (ov === 'auto' || ov === 'scroll')) {
                        el.scrollTop = 0;
                        break;
                    }
                    el = el.parentElement;
                }
            }
        ''')
    except Exception:
        pass
    time.sleep(_human_delay(0.5, 0.3))

    # 检查顶部位置
    try:
        card = target.ele(card_css, timeout=1)
        if card:
            return card
    except Exception:
        pass

    # Phase 3: 从顶部向下逐步滚动搜索
    prev_scroll = -1
    for _ in range(max_scrolls):
        if stop_event and stop_event.is_set():
            return None

        try:
            target.run_js(f'''
                window.scrollBy(0, {scroll_px});
                var card = document.querySelector('[data-geekid]');
                if (card) {{
                    var el = card.parentElement;
                    while (el && el !== document.body && el !== document.documentElement) {{
                        var ov = getComputedStyle(el).overflowY;
                        if (el.scrollHeight > el.clientHeight + 10
                            && (ov === 'auto' || ov === 'scroll')) {{
                            el.scrollTop += {scroll_px};
                            break;
                        }}
                        el = el.parentElement;
                    }}
                }}
            ''')
        except Exception:
            break
        time.sleep(_human_delay(0.3, 0.2))

        try:
            card = target.ele(card_css, timeout=0.5)
            if card:
                return card
        except Exception:
            pass

        # 检测是否已到底（scrollTop 不再变化）
        try:
            cur_scroll = target.run_js('return document.documentElement.scrollTop || document.body.scrollTop || 0')
            if cur_scroll == prev_scroll:
                break
            prev_scroll = cur_scroll
        except Exception:
            pass

    return None


def _parse_greet_context_from_detail_url(detail_url: str) -> dict[str, Any]:
    """Build a persisted greet_context from a BOSS candidate detail API URL."""
    if not detail_url or "/wapi/zpjob/view/geek/info" not in detail_url:
        return {}
    parsed = urlparse(detail_url)
    params = parse_qs(parsed.query, keep_blank_values=True)

    def first(name: str) -> str:
        values = params.get(name) or []
        return str(values[0]) if values else ""

    jid = first("encryptJid") or first("encryptExpectId")
    expect_id = first("expectId")
    security_id = first("securityId")
    lid = first("lid")
    if not jid or not security_id or not lid:
        return {}

    detail_api = {
        "endpoint": parsed.path,
        "encryptJid": first("encryptJid"),
        "encryptExpectId": first("encryptExpectId"),
        "expectId": expect_id,
        "securityId": security_id,
        "lid": lid,
        "entrance": first("entrance"),
        "wayType": first("wayType"),
        "sourceType": first("sourceType"),
    }
    chat_start = {
        "jid": jid,
        "expectId": expect_id,
        "lid": lid,
        "securityId": security_id,
        "greet": "",
        "customGreetingGuide": "-1",
    }
    return {
        "version": GREET_CONTEXT_VERSION,
        "captured_at": datetime.now().isoformat(timespec="seconds"),
        "source": "detail_api",
        "detail_api": {k: v for k, v in detail_api.items() if v not in ("", None)},
        "chat_start": chat_start,
    }


def _extract_detail_url_from_packet(packet: Any) -> str:
    """Best-effort URL extraction for DrissionPage listener packets."""
    try:
        url = getattr(packet, "url", "") or ""
        if url:
            return str(url)
    except Exception:
        pass
    try:
        request = getattr(packet, "request", None)
        return str(getattr(request, "url", "") or "")
    except Exception:
        return ""


def _recent_detail_api_urls(*targets: Any) -> list[str]:
    """Read recent candidate detail API URLs from browser performance entries."""
    urls: list[str] = []
    seen: set[str] = set()
    js = r'''
        return (function() {
            try {
                return JSON.stringify(
                    performance.getEntriesByType('resource')
                        .map(e => e && e.name || '')
                        .filter(Boolean)
                        .filter(u => u.indexOf('/wapi/zpjob/view/geek/info') >= 0)
                        .slice(-50)
                );
            } catch (e) {
                return '[]';
            }
        })()
    '''
    expanded_targets: list[Any] = []
    for target in targets:
        if not target:
            continue
        expanded_targets.append(target)
        try:
            expanded_targets.extend(target.eles("tag:iframe"))
        except Exception:
            pass
    for target in expanded_targets:
        try:
            raw = target.run_js(js) or "[]"
            for url in json.loads(raw):
                if url and url not in seen:
                    urls.append(url)
                    seen.add(url)
        except Exception:
            continue
    return urls


def _clear_resource_timings(*targets: Any) -> None:
    """Clear browser resource timing buffers before one detail-capture attempt."""
    expanded_targets: list[Any] = []
    for target in targets:
        if not target:
            continue
        expanded_targets.append(target)
        try:
            expanded_targets.extend(target.eles("tag:iframe"))
        except Exception:
            pass
    for target in expanded_targets:
        try:
            target.run_js("try { performance.clearResourceTimings(); } catch (e) {}", timeout=1)
        except Exception:
            continue


def _click_candidate_card_for_detail(target: Any, geek_id: str) -> bool:
    """Open a candidate detail panel by clicking the card body, avoiding action buttons."""
    script = r'''
        const card = document.querySelector('[data-geekid="' + arguments[0] + '"]');
        if (!card) return false;
        card.scrollIntoView({block: 'center', inline: 'nearest'});
        const blocked = /(立即沟通|继续沟通|打招呼|聊天|联系|收藏|点赞|分享)/;
        let opener = null;
        const candidates = Array.from(card.querySelectorAll('*')).filter(el => {
            const text = (el.innerText || '').replace(/\s+/g, ' ').trim();
            const tag = (el.tagName || '').toLowerCase();
            if (!text || blocked.test(text)) return false;
            if (tag === 'button' || tag === 'a') return false;
            const rect = el.getBoundingClientRect();
            return rect.width >= 40 && rect.height >= 16;
        });
        opener = candidates.find(el => /(name|info|content|title|text)/i.test(el.className || ''))
              || candidates[0]
              || card;
        if (typeof opener.click === 'function') {
            opener.click();
        } else {
            opener.dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true, view: window}));
        }
        return true;
    '''
    try:
        return bool(target.run_js(script, str(geek_id)))
    except Exception:
        return False


def _close_detail_drawer(page: ChromiumPage) -> None:
    """Close the candidate detail drawer after greet_context capture.

    Two ESC strategies: (1) real browser-level key via CDP Actions, which
    fires through the native event path and is more reliable than synthetic
    events; (2) synthetic KeyboardEvent on document as fallback.
    """
    # Strategy 1: real browser-level ESC via CDP Actions.
    try:
        page.actions.key_down('Escape')
        page.actions.key_up('Escape')
        time.sleep(0.3)
    except Exception:
        pass

    # Strategy 2: synthetic ESC on document (original behaviour, fallback).
    try:
        iframe = get_iframe(page) or page
        iframe.run_js(
            "document.dispatchEvent(new KeyboardEvent('keydown', "
            "{key:'Escape', code:'Escape', keyCode:27, bubbles:true}))"
        )
    except Exception:
        pass


def _capture_greet_context_from_list_page(
    page: ChromiumPage,
    geek_id: str,
    stop_event=None,
    timeout: float = 4.0,
) -> tuple[dict[str, Any], str]:
    """Open candidate detail from the recommendation list and capture greet_context.

    This is a best-effort enrichment step. It must not be treated as required for
    scanning or filtering because BOSS page structure and account state can drift.
    """
    if stop_event and stop_event.is_set():
        return {}, "已停止"
    try:
        iframe = get_iframe(page)
        target = iframe if iframe else page
        card_css = _sel('candidate_card', 'card_by_id_css',
                        'css:[data-geekid="{geek_id}"]').format(geek_id=geek_id)
        card = target.ele(card_css, timeout=1)
        if not card:
            card = _find_card_by_scroll(target, card_css, stop_event=stop_event, max_scrolls=8)
        if not card:
            return {}, "未找到候选人卡片"

        _clear_resource_timings(page)

        if not _click_candidate_card_for_detail(target, geek_id):
            return {}, "打开详情失败"

        deadline = time.time() + timeout
        while time.time() < deadline:
            if stop_event and stop_event.is_set():
                break
            for url in _recent_detail_api_urls(page):
                context = _parse_greet_context_from_detail_url(url)
                if context:
                    return context, "成功"
            time.sleep(0.25)
        return {}, "未捕获详情接口"
    except Exception as exc:
        return {}, f"异常: {str(exc)[:50]}"
    finally:
        try:
            _close_detail_drawer(page)
        except Exception:
            pass


def enrich_greet_contexts_for_candidates(
    page: ChromiumPage,
    candidates: list[dict[str, Any]],
    stop_event=None,
    max_count: int | None = None,
) -> int:
    """Best-effort enrichment of saved greet_context for candidate records."""
    enriched = 0
    attempted = 0
    for candidate in candidates:
        if stop_event and stop_event.is_set():
            raise StopRequested()
        geek_id = candidate.get("geek_id")
        if not geek_id:
            continue
        if max_count is not None and attempted >= max_count:
            break
        attempted += 1
        if attempted > 1 and (attempted - 1) % GREET_CONTEXT_BATCH_SIZE == 0:
            pause = _human_delay(GREET_CONTEXT_BATCH_PAUSE_CENTER, GREET_CONTEXT_BATCH_PAUSE_SPREAD)
            print(f"  已补抓 {attempted - 1} 人上下文，暂停 {int(pause)} 秒降低访问频率...")
            time.sleep(pause)
        had_context = bool(candidate.get("greet_context"))
        context, msg = _capture_greet_context_from_list_page(page, str(geek_id), stop_event=stop_event)
        if context:
            candidate["greet_context"] = context
            candidate["greet_context_updated_at"] = datetime.now().isoformat(timespec="seconds")
            enriched += 1
            action = "已刷新" if had_context else "已保存"
            print(f"  {action} {candidate.get('name', '候选人')} 的打招呼上下文")
        else:
            print(f"  未更新 {candidate.get('name', '候选人')} 的打招呼上下文：{msg}")
        time.sleep(_human_delay(GREET_CONTEXT_DELAY_CENTER, GREET_CONTEXT_DELAY_SPREAD))
    return enriched


def _ensure_zhipin_origin_for_fetch(page: ChromiumPage) -> tuple[bool, str]:
    """Ensure browser JavaScript runs under zhipin.com so same-origin fetch works."""
    try:
        current_url = str(getattr(page, "url", "") or "")
    except Exception:
        current_url = ""
    if "zhipin.com" in current_url.lower():
        return True, ""
    try:
        page.get("https://www.zhipin.com/")
        time.sleep(_human_delay(0.6, 0.3))
        current_url = str(getattr(page, "url", "") or "")
        if "zhipin.com" in current_url.lower():
            return True, ""
        return False, f"无法打开 BOSS 同源页面: {current_url or '未知页面'}"
    except Exception as exc:
        return False, f"无法打开 BOSS 同源页面: {str(exc)[:50]}"


def send_greeting_with_context(
    page: ChromiumPage,
    greet_context: dict[str, Any],
    stop_event=None,
    captcha_callback=None,
) -> tuple[bool, str]:
    """Send a detail-context greeting through /wapi/zpjob/chat/start."""
    if stop_event and stop_event.is_set():
        return False, "已停止"
    chat_start = (greet_context or {}).get("chat_start") or {}
    required = ("jid", "lid", "securityId")
    missing = [key for key in required if not chat_start.get(key)]
    if missing:
        return False, f"缺少打招呼上下文字段: {', '.join(missing)}"

    origin_ok, origin_msg = _ensure_zhipin_origin_for_fetch(page)
    if not origin_ok:
        return False, origin_msg

    payload = {
        "jid": str(chat_start.get("jid") or ""),
        "expectId": str(chat_start.get("expectId") or ""),
        "lid": str(chat_start.get("lid") or ""),
        "greet": str(chat_start.get("greet") or ""),
        "securityId": str(chat_start.get("securityId") or ""),
        "customGreetingGuide": str(chat_start.get("customGreetingGuide") or "-1"),
    }
    payload = {k: v for k, v in payload.items() if v != "" or k in {"greet", "customGreetingGuide"}}
    body = urlencode(payload)
    script = r'''
        var xhr = new XMLHttpRequest();
        xhr.open('POST', '/wapi/zpjob/chat/start', false);
        xhr.withCredentials = true;
        xhr.setRequestHeader('content-type', 'application/x-www-form-urlencoded; charset=UTF-8');
        xhr.setRequestHeader('x-requested-with', 'XMLHttpRequest');
        try {
            xhr.send(arguments[0]);
            var payload = null;
            try { payload = JSON.parse(xhr.responseText || '{}'); } catch (e) { payload = {}; }
            return JSON.stringify({
                ok: xhr.status >= 200 && xhr.status < 300,
                status: xhr.status,
                code: payload && payload.code,
                message: payload && (payload.message || payload.msg),
                zpDataKeys: payload && payload.zpData && typeof payload.zpData === 'object'
                    ? Object.keys(payload.zpData).slice(0, 50)
                    : []
            });
        } catch (err) {
            return JSON.stringify({ok:false, status:xhr.status || 0, error:String(err)});
        }
    '''
    try:
        raw = page.run_js(script, body, timeout=15)
        result = json.loads(raw or "{}") if isinstance(raw, str) else (raw or {})
    except Exception as exc:
        return False, f"上下文打招呼异常: {str(exc)[:50]}"

    if result.get("status") == 200 and result.get("code") == 0:
        return True, "成功"

    is_limited, limit_msg = _detect_limit_popup(page)
    is_captcha, captcha_msg = _detect_captcha(page)
    if is_limited:
        return False, f"沟通次数已达上限: {limit_msg}"
    if is_captcha:
        print(f"\n   打招呼时检测到安全验证弹窗 ({captcha_msg})")
        if _wait_for_captcha_resolution(page, stop_event, captcha_callback=captcha_callback, detail=captcha_msg, stage="greeting"):
            return False, "安全验证已完成，请重新发起本次手工打招呼"
        return False, f"安全验证未完成: {captcha_msg}"

    message = result.get("message") or result.get("error") or "未知响应"
    return False, f"上下文打招呼失败: HTTP {result.get('status')} code={result.get('code')} {message}"


def send_greeting_on_list_page(
    page,
    geek_id,
    retry=0,
    stop_event=None,
    captcha_callback=None,
) -> tuple[bool | None, str]:
    """
    在列表页直接向候选人打招呼（极速优化版）

    优化要点：
    - 智能滚动搜索替代单次 300px 滚动，系统性定位不在可视区的卡片
    - CSS 选择器替代 //* 全局 XPath 扫描
    - 合并按钮文本查询为单次 XPath OR 表达式，消灭循环等待叠加
    - 所有 ele() 调用设短超时，不再死等默认 10s
    - 点击后检测升级套餐/次数上限弹窗，防止假成功
    - 检测到安全验证弹窗时暂停等待用户处理，验证完成后自动重试

    返回：(是否成功，消息)
    """
    try:
        iframe = get_iframe(page)
        target = iframe if iframe else page

        # CSS 选择器按 data-geekid 属性定位卡片，比 //* 快几个数量级
        card_css = _sel('candidate_card', 'card_by_id_css',
                        'css:[data-geekid="{geek_id}"]').format(geek_id=geek_id)
        card = target.ele(card_css, timeout=2)

        if not card:
            # 智能滚动搜索：当前位置 → 滚到顶部 → 逐步向下（最多 40 轮 × 800px）
            card = _find_card_by_scroll(target, card_css, stop_event=stop_event)

        if not card:
            return False, "未找到卡片(滚动搜索后仍不在可视区)"

        parent = card.parent()
        if not parent:
            return False, "未找到卡片父容器"

        # 合并 XPath：单次查询匹配三种按钮文本，彻底消灭 for 循环 + 默认超时叠加
        xpath_query = _sel('greet_button', 'button_xpath',
            'xpath:.//*[text()="继续沟通" or text()="立即沟通" or text()="打招呼" '
            'or contains(text(), "继续沟通") or contains(text(), "立即沟通") '
            'or contains(text(), "打招呼")]')
        greet_btn = parent.ele(xpath_query, timeout=2)

        if not greet_btn:
            return False, "未找到按钮"

        before_button_text = str(getattr(greet_btn, "text", "") or "").strip()
        continue_mark = _sel('greeting_verify', 'continue_mark', "继续沟通")
        if continue_mark in before_button_text:
            return True, "BOSS 页面已显示“继续沟通”，确认此前已建立沟通"

        # 滚到可见区域再点击，防止被悬浮头部遮挡
        try:
            greet_btn.scroll.to_see(center=True)
            time.sleep(_human_delay(0.1, 0.08))
            greet_btn.click()
        except Exception:
            greet_btn.run_js('this.click()')

        time.sleep(_human_delay(0.3, 0.3))  # 等待按钮点击生效

        # 检测升级套餐/次数上限弹窗（重试一次，弹窗可能有渲染延迟）
        is_limited, limit_msg = _detect_limit_popup(page)
        is_captcha, captcha_msg = _detect_captcha(page)
        if not is_limited and not is_captcha:
            time.sleep(_human_delay(0.3, 0.3))
            is_limited, limit_msg = _detect_limit_popup(page)
            if not is_captcha:
                is_captcha, captcha_msg = _detect_captcha(page)

        if is_limited:
            return False, f"沟通次数已达上限: {limit_msg}"
        if is_captcha:
            print(f"\n   打招呼时检测到安全验证弹窗 ({captcha_msg})")
            if _wait_for_captcha_resolution(page, stop_event, captcha_callback=captcha_callback, detail=captcha_msg, stage="greeting"):
                # 验证完成后重新检测弹窗状态。验证码出现说明当前账号/环境已被风控关注，
                # 不继续自动点击下一批，避免刚解除验证又触发更严格的拦截。
                time.sleep(_human_delay(0.5, 0.3))
                is_limited, limit_msg = _detect_limit_popup(page)
                is_captcha, captcha_msg = _detect_captcha(page)
                if is_limited:
                    return False, f"沟通次数已达上限: {limit_msg}"
                if is_captcha:
                    return False, f"验证后仍存在安全弹窗: {captcha_msg}"
                return False, "安全验证已完成，为降低风控风险已停止本轮自动打招呼"
            else:
                return False, f"安全验证未完成: {captcha_msg}"

        return verify_greeting_success(
            target,
            geek_id,
            before_button_text=before_button_text,
            stop_event=stop_event,
        )

    except Exception as e:
        return False, f"异常: {str(e)[:50]}"



def _detect_limit_popup(page: ChromiumPage) -> tuple[bool, str]:
    """
    检测是否弹出了 BOSS 直聘沟通次数上限/升级套餐弹窗。

    明确的“已用完/已达上限”文案可直接判定；“升级套餐”等宽泛文案
    必须出现在可见对话框内并同时带有沟通次数语境，避免把页面上的
    “今日剩余 2 次”等正常额度提示误判为次数耗尽。

    返回: (is_limited: bool, detail: str)
    """
    exhausted_keywords = _sel('limit_detection', 'exhausted_keywords', [
        "次数已用完", "沟通次数已达上限", "联系次数已达上限",
        "今日沟通次数已达上限", "免费次数已用完", "次数不足",
    ])
    upgrade_keywords = _sel('limit_detection', 'upgrade_keywords', [
        "升级套餐", "立即升级", "开通套餐", "购买套餐",
        "升级VIP", "VIP无限沟通", "体验VIP",
    ])
    quota_keywords = _sel('limit_detection', 'quota_keywords', [
        "沟通次数", "联系次数", "免费次数", "今日上限", "今日剩余", "今日免费",
    ])
    config_json = json.dumps({
        "exhausted": exhausted_keywords,
        "upgrade": upgrade_keywords,
        "quota": quota_keywords,
    }, ensure_ascii=False)
    script = (
        'return (function(){'
        f'var cfg={config_json};'
        'function vis(el){'
        'if(!el)return false;var n=el;'
        'while(n&&n.nodeType===1){var s=getComputedStyle(n);'
        'if(s.display==="none"||s.visibility==="hidden"||s.opacity==="0")return false;'
        'n=n.parentElement;}'
        'var r=el.getBoundingClientRect();'
        'return r.width>=10&&r.height>=10&&r.bottom>=0&&r.top<=window.innerHeight'
        '&&r.right>=0&&r.left<=window.innerWidth;}'
        'function hit(text,kws){for(var i=0;i<kws.length;i++){'
        'if(text.indexOf(kws[i])!==-1)return kws[i];}return "";}'
        'var visibleText="",tw=document.createTreeWalker(document.body,4,null),nd;'
        'while(nd=tw.nextNode()){if(nd.parentElement&&vis(nd.parentElement))'
        'visibleText+="\\n"+nd.textContent;}'
        'var exhausted=hit(visibleText,cfg.exhausted);'
        'if(exhausted)return JSON.stringify({matched:exhausted,scope:"visible page"});'
        'var selectors=["[role=dialog]","[aria-modal=true]",'
        '"[class*=dialog]","[class*=modal]","[class*=popup]"];'
        'var nodes=document.querySelectorAll(selectors.join(","));'
        'for(var i=0;i<nodes.length;i++){if(!vis(nodes[i]))continue;'
        'var text=nodes[i].innerText||"";'
        'var upgrade=hit(text,cfg.upgrade),quota=hit(text,cfg.quota);'
        'if(upgrade&&quota)return JSON.stringify({matched:upgrade+" + "+quota,scope:"dialog"});}'
        'return "";})()'
    )

    try:
        # 优先检查 iframe（BOSS 弹窗通常在此渲染），主页面作为兜底。
        iframe = get_iframe(page)
        if iframe:
            try:
                matched = iframe.run_js(script)
                if matched:
                    detail = json.loads(matched)
                    return True, f"iframe 检测到限制提示（匹配：{detail['matched']}）"
            except Exception:
                pass

        try:
            matched = page.run_js(script)
            if matched:
                detail = json.loads(matched)
                return True, f"主页面检测到限制提示（匹配：{detail['matched']}）"
        except Exception:
            pass

        return False, ""

    except Exception:
        return False, ""


def _detect_captcha(page: ChromiumPage) -> tuple[bool, str]:
    """
    检测是否弹出了 BOSS 直聘安全验证弹窗（滑块/图形验证码等）

    BOSS 直聘常见验证形式：
        - 滑块拼图验证（.captcha-slider, .slider-verify 等）
        - 图形验证码（.geetest, .captcha-img 等）
        - 安全提醒弹窗（"检测到异常操作"、"请完成安全验证"）

    检测到验证码时暂停自动化并提示用户处理，避免盲目重试触发风控升级。

    返回: (is_captcha: bool, detail: str)
    """
    captcha_keywords = _sel('captcha_detection', 'keywords', [
        "请完成安全验证", "滑块验证", "请拖动滑块",
        "拖拽拼图", "检测到异常操作", "请先完成验证",
        "行为验证", "请完成验证", "拖动下方滑块", "请按住滑块",
    ])

    # 构建只检查可见文本节点的 JS（排除 display:none / visibility:hidden / 视口外元素）
    # \\n in Python → \n in JS (literal newline inside JS string)
    kw_js = "\\n".join(captcha_keywords)
    script = (
        'return (function(){'
        'var kws=("' + kw_js + '").split("\\n");'
        'function vis(el){'
        'var n=el;'
        'while(n&&n.nodeType===1){'
        'var s=getComputedStyle(n);'
        'if(s.display==="none"||s.visibility==="hidden"||s.opacity==="0")return false;'
        'n=n.parentElement;}'
        'var r=el.getBoundingClientRect();'
        'if(r.width<10||r.height<10)return false;'
        'var vw=window.innerWidth,vh=window.innerHeight;'
        'if(r.bottom<0||r.top>vh||r.right<0||r.left>vw)return false;'
        'return true;}'
        'var tw=document.createTreeWalker(document.body,4,null);'
        'var nd;'
        'while(nd=tw.nextNode()){'
        'if(nd.parentElement&&vis(nd.parentElement)){'
        'for(var i=0;i<kws.length;i++){if(nd.textContent.indexOf(kws[i])!==-1)return kws[i];}}}'
        'return "";})()'
    )

    try:
        iframe = get_iframe(page)
        if iframe:
            try:
                matched_kw = iframe.run_js(script)
                if matched_kw:
                    return True, f"iframe 检测到安全验证弹窗（匹配词：{matched_kw}）"
            except Exception:
                pass

        try:
            matched_kw = page.run_js(script)
            if matched_kw:
                return True, f"主页面检测到安全验证弹窗（匹配词：{matched_kw}）"
        except Exception:
            pass

        # 额外检查常见验证码容器的 DOM 存在性（无文本提示的纯图形验证）
        container_checks = _sel('captcha_detection', 'css_selectors', [
            '.geetest_panel', '.captcha-box', '.slider-captcha',
            '.verify-captcha', '.captcha-container', '#captcha',
            '.yoda-modal',
        ])
        for selector in container_checks:
            try:
                el = page.ele(f'css:{selector}', timeout=0.3)
                if el and el.states.is_displayed:
                    return True, f"检测到验证码容器 ({selector})"
            except Exception:
                continue

        return False, ""

    except Exception:
        return False, ""


def _collect_captcha_diagnostic(page: ChromiumPage, detail: str = "", stage: str = "") -> Path | None:
    """保存验证码现场诊断信息，供用户反馈不同弹窗形态时定位。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stage = re.sub(r"[^a-zA-Z0-9_-]+", "_", stage or "unknown").strip("_") or "unknown"
    out_dir = BASE_DIR / "captcha_diagnostics"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.warning("创建验证码诊断目录失败：%s", e)
        return None

    def _run_js(target: Any, script: str) -> Any:
        try:
            return target.run_js(script)
        except Exception:
            return None

    iframe = None
    try:
        iframe = get_iframe(page)
    except Exception:
        iframe = None
    target = iframe if iframe else page

    href = _run_js(target, "return location.href") or _run_js(page, "return location.href") or getattr(page, "url", "")
    title = _run_js(target, "return document.title") or _run_js(page, "return document.title") or ""
    visible_text = _run_js(
        target,
        "return (document.body && document.body.innerText || '').replace(/\\s+/g, ' ').trim().slice(0, 4000)",
    ) or ""

    screenshot_path = out_dir / f"{timestamp}_{safe_stage}.png"
    screenshot_saved = False
    for method_name, kwargs in (
        ("get_screenshot", {"path": str(screenshot_path)}),
        ("get_screenshot", {"name": str(screenshot_path)}),
        ("save_screenshot", {"path": str(screenshot_path)}),
        ("screenshot", {"path": str(screenshot_path)}),
    ):
        method = getattr(page, method_name, None)
        if not method:
            continue
        try:
            method(**kwargs)
            screenshot_saved = screenshot_path.exists()
            if screenshot_saved:
                break
        except Exception:
            continue

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stage": stage,
        "detail": detail,
        "url": href,
        "title": title,
        "screenshot": screenshot_path.name if screenshot_saved else "",
        "visible_text_excerpt": visible_text,
    }
    json_path = out_dir / f"{timestamp}_{safe_stage}.json"
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"   已保存验证码诊断：{json_path}")
        return json_path
    except Exception as e:
        logger.warning("保存验证码诊断失败：%s", e)
        return None


def _wait_for_captcha_resolution(page, stop_event=None, max_wait=CAPTCHA_MAX_WAIT, captcha_callback=None, detail="", stage=""):
    """
    等待用户手动完成安全验证（验证码/滑块）。

    每隔 3 秒检查验证码是否已消失。用户完成验证后自动恢复。
    支持 stop_event 中断和最大等待时间。

    参数:
        page: DrissionPage 页面对象
        stop_event: threading.Event，设置时立即返回 False
        max_wait: 最大等待秒数（默认 5 分钟），超时返回 False
        captcha_callback: callable(detail) -> bool，检测到验证码时调用，
            用于 GUI 弹窗通知用户。返回 True 继续等待，False 中止。
            若为 None 则仅通过日志通知。

    返回:
        True: 验证已完成，可继续
        False: 用户中止或超时
    """
    print(f"\n⚠️  检测到安全验证弹窗，请在浏览器中手动完成验证。")
    print(f"   程序将自动检测验证状态，完成后继续运行...（最长等待 {max_wait} 秒）")
    _collect_captcha_diagnostic(page, detail=detail, stage=stage or "captcha")

    # 通知 GUI 弹窗
    if captcha_callback:
        try:
            user_continue = captcha_callback(detail)
            if not user_continue:
                print("   用户选择跳过验证等待。")
                return False
        except Exception:
            pass

    elapsed = 0
    check_interval = CAPTCHA_CHECK_INTERVAL
    while elapsed < max_wait:
        if stop_event and stop_event.is_set():
            print("   用户中止，停止等待验证。")
            return False
        time.sleep(check_interval)
        elapsed += check_interval
        is_still, _ = _detect_captcha(page)
        if not is_still:
            print(f"✅ 验证已完成，继续运行...")
            return True
        if elapsed % 15 == 0:
            remaining = max_wait - elapsed
            print(f"   仍在等待验证...（剩余 {remaining} 秒，可随时在浏览器中完成验证）")

    print(f"⏰ 等待验证超时（{max_wait} 秒），停止当前操作。")
    return False


def verify_greeting_success(
    target: Any,
    geek_id: str,
    *,
    before_button_text: str = "",
    stop_event=None,
    attempts: int = 5,
    interval: float = 0.8,
) -> tuple[bool | None, str]:
    """
    验证列表页打招呼结果。

    True 表示页面出现明确成功状态；False 仅保留给明确失败；
    None 表示点击已执行但页面状态无法确认，调用方不得落盘为已沟通。
    """
    card_css = _sel(
        'candidate_card',
        'card_by_id_css',
        'css:[data-geekid="{geek_id}"]',
    ).format(geek_id=geek_id)
    success_marks = _sel(
        'greeting_verify',
        'success_marks',
        ["已沟通", "沟通过", "已发送"],
    )
    continue_mark = _sel('greeting_verify', 'continue_mark', "继续沟通")
    last_state = "候选人卡片未重新出现"

    for attempt in range(max(1, attempts)):
        if stop_event and stop_event.is_set():
            return None, "点击已执行，但用户停止后未完成发送结果确认"
        if attempt:
            time.sleep(interval)
        try:
            card = target.ele(card_css, timeout=0.5)
            if not card:
                last_state = "候选人卡片未重新出现"
                continue
            parent = card.parent()
            if not parent:
                last_state = "候选人卡片父容器不可用"
                continue
            all_text = str(getattr(parent, "text", "") or "")
            matched = next((mark for mark in success_marks if mark in all_text), "")
            if matched:
                return True, f"发送成功，页面出现“{matched}”"
            if continue_mark in all_text:
                if continue_mark in before_button_text:
                    return True, "BOSS 页面已显示“继续沟通”，确认此前已建立沟通"
                return True, "发送成功，按钮已变为“继续沟通”"
            last_state = f"按钮尚未变为“{continue_mark}”"
        except Exception as exc:
            last_state = f"页面状态读取异常：{str(exc)[:50]}"

    return None, f"点击已执行，但发送结果无法确认（{last_state}），请在 BOSS 沟通列表核实"


def check_selectors_health(page: ChromiumPage) -> list[dict[str, Any]]:
    """选择器健康检查：逐一测试 selectors.json 中的关键选择器，返回诊断报告。

    返回: list[dict]，每项包含:
        - group: 选择器组名
        - name: 选择器名称
        - status: 'ok' | 'warn' | 'fail'
        - detail: 描述信息
    """
    # Chrome 刷新或重建标签页时，DrissionPage 的旧对象可能短暂断线。
    # 先验证连接，避免把连接问题误报成选择器失效。
    page.run_js('return 1')
    results = []
    sel = load_selectors()

    def _check(name, test_fn, group="", expect_found=True):
        try:
            found = test_fn()
            if expect_found and found:
                results.append({'group': group, 'name': name, 'status': 'ok',
                               'detail': f'找到 {found} 个匹配元素'})
            elif expect_found and not found:
                results.append({'group': group, 'name': name, 'status': 'warn',
                               'detail': '未找到匹配元素（可能页面未加载或选择器已失效）'})
            else:
                results.append({'group': group, 'name': name, 'status': 'ok',
                               'detail': '检查通过'})
        except Exception as e:
            results.append({'group': group, 'name': name, 'status': 'fail',
                           'detail': f'异常：{type(e).__name__}: {str(e)[:80]}'})

    # 1. iframe 检测
    iframe_sel = _sel('iframe', 'selector', 'tag:iframe')
    def _test_iframe():
        frames = page.eles(iframe_sel)
        return len(frames) if frames else 0
    _check('iframe', _test_iframe, group='iframe', expect_found=False)

    # 2. 候选人卡片
    cards_sel = _sel('candidate_card', 'all_cards_xpath', 'xpath://*[@data-geekid]')
    iframe = get_iframe(page)
    target = iframe if iframe else page
    def _test_cards():
        els = target.eles(cards_sel)
        return len(els) if els else 0
    _check('all_cards', _test_cards, group='candidate_card')

    # 3. 验证码容器（不应存在）
    captcha_css = _sel('captcha_detection', 'css_selectors', [])
    captcha_found = []
    for css in captcha_css:
        try:
            el = page.ele(f'css:{css}', timeout=0.3)
            if el and el.states.is_displayed:
                captcha_found.append(css)
        except Exception:
            continue
    if captcha_found:
        results.append({'group': 'captcha_detection', 'name': 'css_containers',
                       'status': 'warn', 'detail': f'检测到验证码容器: {captcha_found}'})
    else:
        results.append({'group': 'captcha_detection', 'name': 'css_containers',
                       'status': 'ok', 'detail': '无验证码弹窗'})

    # 4. 明确的次数耗尽文案
    limit_kws = _sel('limit_detection', 'exhausted_keywords', [])
    if limit_kws:
        checks = " || ".join(f'body.innerText.includes("{kw}")' for kw in limit_kws)
        script = f'return (function(){{var body=document.body;return {checks};}})()'
        try:
            triggered = target.run_js(script)
            if triggered:
                results.append({'group': 'limit_detection', 'name': 'keywords',
                               'status': 'warn', 'detail': '检测到明确的次数耗尽文案'})
            else:
                results.append({'group': 'limit_detection', 'name': 'keywords',
                               'status': 'ok', 'detail': '无限制弹窗'})
        except Exception as e:
            results.append({'group': 'limit_detection', 'name': 'keywords',
                           'status': 'fail', 'detail': f'JS 执行失败：{str(e)[:60]}'})

    # 5. 打招呼按钮（需要至少有卡片才能检测）
    btn_xpath = _sel('greet_button', 'button_xpath', '')
    if btn_xpath and _test_cards() > 0:
        try:
            cards = target.eles(cards_sel)
            if cards:
                parent = cards[0].parent()
                if parent:
                    btn = parent.ele(btn_xpath, timeout=1)
                    results.append({'group': 'greet_button', 'name': 'button_xpath',
                                   'status': 'ok' if btn else 'warn',
                                   'detail': '找到按钮' if btn else '第一张卡片未找到按钮'})
        except Exception as e:
            results.append({'group': 'greet_button', 'name': 'button_xpath',
                           'status': 'fail', 'detail': str(e)[:80]})

    # 6. selectors.json 文件完整性
    expected_groups = ['candidate_card', 'name_extraction', 'greet_button',
                       'iframe', 'scroll', 'captcha_detection', 'limit_detection']
    missing = [g for g in expected_groups if g not in sel]
    if missing:
        results.append({'group': 'config', 'name': 'selectors.json',
                       'status': 'warn', 'detail': f'缺少配置组: {missing}'})
    else:
        results.append({'group': 'config', 'name': 'selectors.json',
                       'status': 'ok', 'detail': f'配置文件完整 ({len(sel)} 组)'})

    return results

def _select_greet_context_candidates(
    passed_candidates: list[dict[str, Any]],
    *,
    auto_greet: bool,
    point_to_point_mode: bool,
    greet_names_list: list[str] | None,
    greet_levels_allowed: list[str],
    existing_greeted_ids: set[str],
    raw_order_by_geek_id: dict[str, int],
) -> list[dict[str, Any]]:
    """选择本轮不会自动发送、但后续可能需要手工补发的候选人。"""
    if point_to_point_mode:
        planned_auto_greet = [
            c for c in passed_candidates
            if c.get('name') in (greet_names_list or [])
            and c.get('geek_id') not in existing_greeted_ids
        ]
    else:
        planned_auto_greet = [
            c for c in passed_candidates
            if c.get('recommend_level') in greet_levels_allowed
            and c.get('geek_id') not in existing_greeted_ids
            and not c.get('manual_review_required')
        ]
    planned_auto_greet.sort(
        key=lambda x: raw_order_by_geek_id.get(
            str(x.get('geek_id')), len(raw_order_by_geek_id)
        )
    )
    planned_ids = {
        c.get('geek_id')
        for c in planned_auto_greet[:AUTO_GREET_RUN_LIMIT]
    } if auto_greet else set()

    return [
        c for c in passed_candidates
        if c.get('geek_id') not in existing_greeted_ids
        and c.get('geek_id') not in planned_ids
        and c.get('match_score', 0) >= GREET_CONTEXT_MIN_SCORE
        and c.get('qualification_status') != 'rejected'
    ]


def _prioritize_greet_context_candidates(
    candidates: list[dict[str, Any]],
    raw_order_by_geek_id: dict[str, int],
    limit: int = GREET_CONTEXT_CAPTURE_LIMIT,
) -> list[dict[str, Any]]:
    """按业务价值选名额，再按页面顺序返回实际执行队列。"""
    ranked = sorted(candidates, key=lambda c: (
        bool((c.get('greet_context') or {}).get('chat_start')),
        -float(c.get('match_score', 0) or 0),
        raw_order_by_geek_id.get(
            str(c.get('geek_id')), len(raw_order_by_geek_id)
        ),
    ))
    selected = ranked[:limit]
    return sorted(
        selected,
        key=lambda c: raw_order_by_geek_id.get(
            str(c.get('geek_id')), len(raw_order_by_geek_id)
        ),
    )


def smart_scan_candidates(page, job_info, auto_greet=False, max_rounds=MAX_ROUNDS_DEFAULT, verbose=False, greet_level='normal', greet_names_list=None, list_candidates=False, progress_callback=None, stop_event=None, ai_eval=False, api_config=None, api_key=None, captcha_callback=None, notice_callback=None, blocking_notice_callback=None, stats=None, max_candidates=API_CANDIDATE_LIMIT_DEFAULT, use_api_extraction=True, extraction_mode=None):
    """
    智能扫描候选人 - 两阶段模式

    阶段 1: 滚动收集所有候选人并筛选（不打招呼）
    阶段 1.5: AI 辅助评估（可选，对通过筛选的候选人进行 LLM 二次评分）
    阶段 2: 按分数从高到低依次打招呼

    Args:
        page: 浏览器页面对象
        job_info: 岗位信息
        auto_greet: 是否自动打招呼
        max_rounds: 最大滚动轮次
        verbose: 是否输出详细评分信息
        greet_level: 打招呼等级 'strong'=仅强烈推荐，'normal'=强烈推荐 + 推荐
        greet_names_list: 点对点打招呼的姓名列表（如果指定，只给这些候选人打招呼）
        list_candidates: 是否仅列出候选人，不打招呼
        progress_callback: 进度回调 callable(percentage, description)，percentage 0-100
        stop_event: threading.Event，设位时立即停止
        ai_eval: 是否启用 AI 辅助评估
        api_config: API 配置字典（base_url, model），AI 评估时使用
        api_key: API Key 字符串，AI 评估时使用
        captcha_callback: callable(detail) -> bool，检测到验证码时调用，
            用于 GUI 弹窗通知用户。返回 True 继续等待，False 中止。
        notice_callback: callable(title, message)，用于 GUI 展示非阻塞提示。
    """
    job_name = job_info['job_name']

    # 确定打招呼等级的显示和筛选逻辑
    if greet_level == 'strong':
        greet_levels_allowed = ['强烈推荐']
        greet_level_text = '仅强烈推荐'
    else:
        greet_levels_allowed = ['强烈推荐', '推荐']
        greet_level_text = '强烈推荐 + 推荐'

    # 点对点模式
    point_to_point_mode = bool(greet_names_list)

    if point_to_point_mode:
        print(f"开始智能扫描候选人... (岗位：{job_name}, 点对点打招呼：{greet_names_list}, 轮次：{max_rounds})")
    else:
        print(f"开始智能扫描候选人... (岗位：{job_name}, 自动打招呼：{'是 (' + greet_level_text + ')' if auto_greet else '否'}, 轮次：{max_rounds})")

    # 加载 candidates_all.json，检查已打招呼的候选人
    candidates_all = load_candidates_all()
    greeted_geek_ids = get_greeted_geek_ids(candidates_all)
    blacklisted_geek_ids = build_blacklist_index(candidates_all)

    # 从 candidates_all 中获取已匹配的 ID（按岗位过滤）
    existing_ids_for_job_and_greeted = set()  # 当前岗位已匹配且打过招呼的 ID（需要过滤）
    all_existing_ids = set()  # 所有岗位已匹配的 ID
    if candidates_all:
        for c in candidates_all:
            all_existing_ids.add(c['geek_id'])
            if c.get('job_name') == job_name and c.get('greet_sent') is True:
                existing_ids_for_job_and_greeted.add(c['geek_id'])

        blacklist_text = f"，{len(blacklisted_geek_ids)} 人已屏蔽" if blacklisted_geek_ids else ""
        print(f"已加载 candidates_all.json：累计 {len(all_existing_ids)} 个候选人，{len(greeted_geek_ids)} 人已打招呼{blacklist_text}")

    # === 阶段 1: 滚动收集所有候选人 ===
    raw_candidates = extract_candidates_by_comprehensive_analysis(page, max_rounds=max_rounds, progress_callback=progress_callback, stop_event=stop_event, captcha_callback=captcha_callback, notice_callback=notice_callback, blocking_notice_callback=blocking_notice_callback, max_candidates=max_candidates, use_api_extraction=use_api_extraction, extraction_mode=extraction_mode)
    print(f"原始提取到 {len(raw_candidates)} 个唯一候选人")
    existing_by_candidate_key = {
        (str(c.get('geek_id')), c.get('job_name', '')): c
        for c in candidates_all
        if c.get('geek_id')
    }

    if blacklisted_geek_ids:
        before_count = len(raw_candidates)
        raw_candidates = [c for c in raw_candidates if str(c.get('geek_id')) not in blacklisted_geek_ids]
        skipped_count = before_count - len(raw_candidates)
        if skipped_count:
            print(f"过滤黑名单候选人：{before_count} -> {len(raw_candidates)} (已屏蔽 {skipped_count} 人)")

    # 过滤当前岗位已打过招呼的候选人；其余候选人进入本轮重新评估。
    if existing_ids_for_job_and_greeted:
        before_count = len(raw_candidates)
        raw_candidates = [c for c in raw_candidates if c['geek_id'] not in existing_ids_for_job_and_greeted]
        print(
            f"过滤当前岗位已打招呼候选人：{before_count} → {len(raw_candidates)} "
            f"（本轮待评估 {len(raw_candidates)} 人）"
        )

    # 筛选所有候选人（暂不打招呼）
    print("\n=== 阶段 1: 筛选候选人 ===")
    passed_candidates = []  # 通过筛选的候选人（含分数）
    failed_reasons = {}

    # 构建淘汰原因的动态描述（基于实际招聘要求）
    rule = job_info['rule']
    exp_requirement = f"经验不足（要求{rule.get('min_exp', 0)}年以上）"
    # 学历要求：优先取 required_conditions 中的统招本科，否则取 edu 字段
    edu_requirement = rule.get('edu', '不限')
    req_conds = rule.get('required_conditions', [])
    for cond in req_conds:
        if isinstance(cond, str) and '统招' in cond:
            edu_requirement = cond
            break
    edu_requirement = f"学历不符/不足（要求{edu_requirement}）"

    # 年龄、地点、薪资、技术条件要求
    max_age = rule.get('max_age')
    age_requirement = f"年龄不符（要求≤{max_age}岁）" if max_age else "年龄不符"
    work_location = rule.get('work_location', '')
    city_requirement = f"地点不符（要求{work_location}）" if work_location else "地点不符"
    salary_max = rule.get('salary_max')
    salary_requirement = f"薪资不匹配（岗位最高{salary_max}K）" if salary_max else "薪资不匹配"
    tech_keywords = [k.get('name', k) if isinstance(k, dict) else k for k in rule.get('keywords', [])]
    tech_requirement = f"技术条件不符"

    for i, candidate in enumerate(raw_candidates):
        if stop_event and stop_event.is_set():
            raise StopRequested()
        passed, score, details = filter_candidate(candidate['summary'], job_info['rule'], candidate.get('structured'))
        if passed and score >= SCORE_THRESHOLD_PASS:
            # 计算推荐等级
            if score >= SCORE_THRESHOLD_STRONG:
                recommend_level = "强烈推荐"
            elif score >= SCORE_THRESHOLD_RECOMMEND:
                recommend_level = "推荐"
            else:
                recommend_level = "待定"

            # 提取结构化字段（薪资、经验、年龄等）
            summary_info = extract_summary_info(candidate['summary'])
            # API 结构化数据优先覆盖 DOM 解析结果
            structured = candidate.get('structured') or {}
            if structured.get('exp_years'):
                summary_info['exp_years'] = str(structured['exp_years'])
            if structured.get('age'):
                summary_info['age'] = str(structured['age'])
            if structured.get('salary_min') and structured.get('salary_max'):
                summary_info['salary'] = f"{structured['salary_min']}-{structured['salary_max']}K"
            elif structured.get('salary_min'):
                summary_info['salary'] = f"{structured['salary_min']}K"
            if structured.get('city'):
                summary_info['city'] = structured['city']
            if structured.get('job_status'):
                summary_info['job_status'] = structured['job_status']
            # 薪资兜底：解析后仍为空则默认面议
            if not summary_info.get('salary'):
                summary_info['salary'] = '面议'

            candidate_record = {
                "geek_id": candidate['geek_id'],
                "name": candidate['name'],
                "summary": candidate['summary'],
                "job_id": job_info['job_id'],
                "job_name": job_name.replace(" ", ""),  # 去除岗位名称中的空格
                "salary": summary_info.get('salary', ''),
                "age": summary_info.get('age', ''),
                "exp_years": summary_info.get('exp_years', ''),
                "education": summary_info.get('education', ''),
                "city": summary_info.get('city') or _extract_city(candidate['summary']),
                "job_status": summary_info.get('job_status', ''),
                "company": summary_info.get('company', ''),
                "match_rule": job_info['rule_key'],
                "match_score": score,
                "skill_matches": details.get('skill_matches', []),
                "skill_match_ratio": f"{details.get('skill_matched_count', 0)}/{details.get('skill_total', 0)}",
                "score_breakdown": details.get('score_breakdown', {}),
                "score_explanation": details.get('score_explanation', []),
                "keyword_evidence": details.get('keyword_evidence', []),
                "risk_flags": details.get('risk_flags', []),
                "manual_review_required": bool(details.get('manual_review_required')),
                "auto_greet_blocked_reason": details.get('auto_greet_blocked_reason', ''),
                "qualification_status": details.get('qualification_status', 'qualified'),
                "qualification_reasons": details.get('qualification_reasons', []),
                "qualification_evidence": details.get('qualification_evidence', []),
                "recommend_level": recommend_level,
                "batch_timestamp": datetime.now().strftime("%Y%m%d_%H%M%S"),
                "followup_status": "未沟通",
                "greet_sent": False
            }
            existing_record = existing_by_candidate_key.get((str(candidate['geek_id']), job_name.replace(" ", "")))
            if existing_record and existing_record.get('greet_context'):
                candidate_record['greet_context'] = existing_record['greet_context']
                if existing_record.get('greet_context_updated_at'):
                    candidate_record['greet_context_updated_at'] = existing_record['greet_context_updated_at']
            if existing_record and existing_record.get('greet_confirmation_pending'):
                candidate_record['greet_confirmation_pending'] = True
                candidate_record['greet_confirmation_reason'] = existing_record.get(
                    'greet_confirmation_reason', ''
                )
                candidate_record['greet_confirmation_updated_at'] = existing_record.get(
                    'greet_confirmation_updated_at', ''
                )
            # 保留 API 结构化数据和画像供后续使用
            if structured:
                candidate_record['structured'] = structured
            if candidate.get('_api_profile'):
                candidate_record['_api_profile'] = candidate['_api_profile']
            passed_candidates.append(candidate_record)

            if verbose:
                print(f"  [{i+1}/{len(raw_candidates)}] {candidate['name']} - {score}分 ({recommend_level})")
        else:
            if passed:
                # 通过了硬性筛选但评分 < 55，按分数段归类
                if score >= 50:
                    reason = "评分不足(50-54分)"
                elif score >= 40:
                    reason = "评分不足(40-49分)"
                elif score >= 30:
                    reason = "评分不足(30-39分)"
                else:
                    reason = "评分不足(<30分)"
            else:
                reason = details.get('reason', '未知')
                # 合并同类淘汰原因
                if '经验不足' in reason:
                    reason = exp_requirement
                elif '学历不足' in reason or '学历不符' in reason:
                    reason = edu_requirement
                elif '年龄不符' in reason or '年龄超限' in reason:
                    reason = age_requirement
                elif '地点不符' in reason or '城市不符' in reason:
                    reason = city_requirement
                elif '薪资不匹配' in reason or '薪资期望过高' in reason:
                    reason = salary_requirement
                elif '技术不匹配' in reason or '必要条件不满足' in reason:
                    reason = tech_requirement
                elif '筛选异常' in reason:
                    reason = '筛选异常'
            failed_reasons[reason] = failed_reasons.get(reason, 0) + 1

        if (i + 1) % 20 == 0:
            print(f"  已筛选 {i + 1}/{len(raw_candidates)} 个，通过 {len(passed_candidates)} 个")
            if progress_callback:
                pct = int((i + 1) / len(raw_candidates) * 100)
                progress_callback(pct, f"正在智能筛选... {i + 1}/{len(raw_candidates)}")

    raw_order_by_geek_id = {
        str(candidate.get('geek_id')): idx
        for idx, candidate in enumerate(raw_candidates)
        if candidate.get('geek_id')
    }

    # 按分数从高到低排序，保持结果展示和 AI 评估优先级不变。
    passed_candidates.sort(key=lambda x: x.get('match_score', 0), reverse=True)

    qualified_count = sum(
        1 for c in passed_candidates if c.get('qualification_status', 'qualified') == 'qualified'
    )
    manual_review_count = sum(
        1 for c in passed_candidates if c.get('qualification_status') == 'manual_review'
    )
    print(f"\n规则筛选完成：候选池 {len(passed_candidates)}/{len(raw_candidates)} 个")
    print(f"  - 明确合格：{qualified_count} 人")
    print(f"  - 待人工确认：{manual_review_count} 人")
    if failed_reasons:
        total_failed = sum(failed_reasons.values())
        print(f"淘汰原因（共 {total_failed} 人）:")
        def _reason_order(item):
            reason = item[0]
            if '经验不足' in reason:
                return (0, -item[1])
            if '学历' in reason:
                return (1, -item[1])
            if '年龄' in reason:
                return (2, -item[1])
            if '地点' in reason:
                return (3, -item[1])
            if '薪资' in reason:
                return (4, -item[1])
            if '技术条件' in reason:
                return (5, -item[1])
            if '评分不足' in reason:
                return (6, -item[1])
            if '筛选异常' in reason:
                return (7, -item[1])
            return (8, -item[1])

        for reason, count in sorted(failed_reasons.items(), key=_reason_order):
            print(f"  - {reason}: {count} 人")
        # 总数校验
        accounted = len(passed_candidates) + total_failed
        if accounted != len(raw_candidates):
            print(f"⚠️  数量不一致：通过({len(passed_candidates)}) + 淘汰({total_failed}) = {accounted} ≠ 原始({len(raw_candidates)})")

    ai_removed_count = 0
    ai_evaluated_count = 0
    rule_pool_count = len(passed_candidates)
    # === 阶段 1.5: AI 辅助评估（可选）===
    if ai_eval and api_config and api_key and passed_candidates:
        from llm_eval import evaluate_batch
        rule = job_info['rule']
        job_requirement = rule.get('original_requirement', '')
        if not job_requirement:
            job_requirement = f"岗位：{job_name}，{rule.get('min_exp', 0)}年经验，{rule.get('edu', '不限')}学历"

        print(f"\n=== AI 辅助评估（共 {len(passed_candidates)} 人）===")
        # 按规则评分降序排列，确保 AI 评估优先处理最有价值的候选人
        passed_candidates.sort(key=lambda x: x.get('match_score', 0), reverse=True)

        # 构建硬条件摘要，供 LLM 评估时参考
        hard_parts = []
        if rule.get('min_exp'):
            hard_parts.append(f"- 经验：≥{rule['min_exp']}年")
        if rule.get('edu') and rule.get('edu') != '不限':
            hard_parts.append(f"- 学历：{rule['edu']}")
        if rule.get('max_age'):
            hard_parts.append(f"- 年龄：≤{rule['max_age']}岁")
        if rule.get('work_location'):
            hard_parts.append(f"- 地点：{rule['work_location']}")
        if rule.get('salary_max'):
            hard_parts.append(f"- 薪资上限：{rule['salary_max']}K")
        req_conds = rule.get('required_conditions', [])
        if req_conds:
            cond_names = [c if isinstance(c, str) else c.get('name', str(c)) for c in req_conds]
            hard_parts.append(f"- 必要条件：{'、'.join(cond_names)}")
        hard_conditions = "## 筛选硬条件\n" + "\n".join(hard_parts) + "\n\n" if hard_parts else ""

        passed_candidates = evaluate_batch(
            passed_candidates, job_requirement, api_config, api_key,
            hard_conditions=hard_conditions,
            progress_callback=progress_callback,
            stop_event=stop_event,
        )
        # 重新排序（分数可能变化）
        passed_candidates.sort(key=lambda x: x.get('match_score', 0), reverse=True)
        llm_count = sum(1 for c in passed_candidates if c.get('llm_evaluated'))
        ai_evaluated_count = llm_count
        llm_failed_count = rule_pool_count - llm_count
        success_rate = llm_count * 100 / rule_pool_count if rule_pool_count else 0
        print(
            f"AI 评估完成：成功 {llm_count} 人，失败 {llm_failed_count} 人，"
            f"成功率 {success_rate:.1f}%"
        )
        if llm_failed_count and llm_failed_count / rule_pool_count >= 0.3:
            print("[WARN] AI 评估失败率过高，本轮结果主要依据规则评分，请人工复核失败候选人")
        ai_hard_rejected = sum(
            1 for c in passed_candidates if c.get('qualification_status') == 'rejected'
        )
        ai_score_rejected = sum(
            1 for c in passed_candidates
            if c.get('qualification_status') != 'rejected'
            and c.get('match_score', 0) < SCORE_THRESHOLD_PASS
        )
        ai_removed_count = ai_hard_rejected + ai_score_rejected
        passed_candidates = [
            c for c in passed_candidates
            if c.get('qualification_status') != 'rejected'
            and c.get('match_score', 0) >= SCORE_THRESHOLD_PASS
        ]
        final_qualified = sum(
            1 for c in passed_candidates if c.get('qualification_status', 'qualified') == 'qualified'
        )
        final_manual = sum(
            1 for c in passed_candidates if c.get('qualification_status') == 'manual_review'
        )
        print("最终评估结果：")
        print(f"  - 明确合格：{final_qualified} 人")
        print(f"  - 待人工确认：{final_manual} 人")
        print(f"  - AI复核硬条件淘汰：{ai_hard_rejected} 人")
        print(f"  - 分数降至{SCORE_THRESHOLD_PASS}分以下：{ai_score_rejected} 人")
        print(f"  - 最终保留：{len(passed_candidates)}/{rule_pool_count} 人")

    # === 阶段 1.6: 为后续可能补打招呼的候选人捕获上下文（最佳努力）===
    # 本轮确定会自动发送的候选人无需提前抓取；被模式、人工确认或单轮上限
    # 排除的人仍保存上下文，方便后续从筛选结果页直接发送。
    context_candidates = _select_greet_context_candidates(
        passed_candidates,
        auto_greet=auto_greet,
        point_to_point_mode=point_to_point_mode,
        greet_names_list=greet_names_list,
        greet_levels_allowed=greet_levels_allowed,
        existing_greeted_ids=existing_ids_for_job_and_greeted,
        raw_order_by_geek_id=raw_order_by_geek_id,
    )
    context_candidates = _prioritize_greet_context_candidates(
        context_candidates,
        raw_order_by_geek_id,
    )
    if context_candidates and not list_candidates and hasattr(page, "listen"):
        print(f"\n=== 阶段 1.6: 捕获打招呼上下文（共 {len(context_candidates)} 人，最佳努力）===")
        enriched_count = enrich_greet_contexts_for_candidates(
            page,
            context_candidates,
            stop_event=stop_event,
            max_count=None,
        )
        if enriched_count:
            merge_candidates_all(context_candidates)
        print(f"打招呼上下文捕获完成：成功 {enriched_count}/{len(context_candidates)} 人")

    # === 仅列出候选人，不打招呼 ===
    if list_candidates and passed_candidates:
        print("\n=== 候选人列表 ===")
        print(f"岗位：{job_name}")
        print(f"总人数：{len(passed_candidates)}")
        print(f"{'序号':<4} {'姓名':<10} {'匹配分':<8} {'推荐指数':<10} {'已打招呼':<8}")
        print("-" * 50)
        for i, c in enumerate(passed_candidates[:50], 1):  # 最多显示前 50 个
            print(f"{i:<4} {c['name']:<10} {c['match_score']:<8} {c['recommend_level']:<10} {'是' if c.get('greet_sent') else '否':<8}")
        if len(passed_candidates) > 50:
            print(f"... 还有 {len(passed_candidates) - 50} 个")
        return passed_candidates

    # === 阶段 2: 按页面顺序依次打招呼 ===
    if auto_greet and passed_candidates:
        print("\n=== 阶段 2: 按页面顺序打招呼 ===")
        print("正在刷新候选人列表...")

        # 温和刷新：小幅度滚动触发懒加载渲染，不滚回顶部破坏虚拟列表
        iframe = get_iframe(page)
        target = iframe if iframe else page
        target.run_js('window.scrollBy(0, 200)')
        time.sleep(_human_delay(0.3, 0.25))
        target.run_js('window.scrollBy(0, -100)')
        time.sleep(_human_delay(0.3, 0.25))

        # 筛选需要打招呼的候选人
        if point_to_point_mode:
            # 点对点模式：只筛选指定姓名的候选人
            to_greet_list = [c for c in passed_candidates
                            if c.get('name') in greet_names_list]
            print(f"需要打招呼：{len(to_greet_list)} 人 (点对点)")
            if not to_greet_list:
                print(f"  未找到匹配的候选人，姓名列表：{greet_names_list}")
        else:
            # 自动模式：根据推荐等级筛选
            to_greet_list = [c for c in passed_candidates
                            if c.get('recommend_level') in greet_levels_allowed
                            and c.get('geek_id') not in existing_ids_for_job_and_greeted
                            and not c.get('greet_confirmation_pending')
                            and not c.get('manual_review_required')]
            blocked_count = sum(
                1 for c in passed_candidates
                if c.get('recommend_level') in greet_levels_allowed
                and c.get('geek_id') not in existing_ids_for_job_and_greeted
                and c.get('manual_review_required')
            )
            pending_count = sum(
                1 for c in passed_candidates
                if c.get('recommend_level') in greet_levels_allowed
                and c.get('geek_id') not in existing_ids_for_job_and_greeted
                and c.get('greet_confirmation_pending')
            )
            print(f"需要打招呼：{len(to_greet_list)} 人 ({greet_level_text})")
            if blocked_count:
                print(f"  已跳过 {blocked_count} 人：需要人工确认后再打招呼")
            if pending_count:
                print(f"  已跳过 {pending_count} 人：上次发送结果待确认，请先在 BOSS 沟通列表核实")

        # 点击顺序按扫描/页面顺序，减少虚拟列表反复回顶和跳跃滚动。
        to_greet_list.sort(key=lambda x: raw_order_by_geek_id.get(str(x.get('geek_id')), len(raw_order_by_geek_id)))
        if len(to_greet_list) > AUTO_GREET_RUN_LIMIT:
            remaining_count = len(to_greet_list) - AUTO_GREET_RUN_LIMIT
            limit_msg = (
                f"为降低 BOSS 风控风险，本轮最多自动打招呼 {AUTO_GREET_RUN_LIMIT} 人，"
                f"剩余 {remaining_count} 人下次继续。\n\n"
                "下次直接再次运行同一岗位扫描即可：程序会跳过当前岗位已打过招呼的人，"
                "未打招呼的合格候选人会重新进入待打招呼队列。"
            )
            print(f"  {limit_msg.replace(chr(10), ' ')}")
            if notice_callback:
                try:
                    notice_callback("已达到本轮自动打招呼上限", limit_msg)
                except Exception:
                    pass
            to_greet_list = to_greet_list[:AUTO_GREET_RUN_LIMIT]

        greet_success_count = 0
        greet_fail_count = 0
        greet_pending_count = 0
        greeted_in_this_run = []
        consecutive_failures = 0  # 连续失败计数
        consecutive_uncertain = 0

        try:
            for i, candidate in enumerate(to_greet_list):
                if stop_event and stop_event.is_set():
                    raise StopRequested()
                if not _ensure_recommend_page(page, notice_callback=notice_callback, context="打招呼"):
                    break
                action = "补打招呼" if candidate['geek_id'] in all_existing_ids else "打招呼"

                # 检查连续失败，如果连续失败 3 次则停止
                if consecutive_failures >= GREET_FAIL_LIMIT:
                    print(f"\n⚠️  连续 {consecutive_failures} 次失败，停止打招呼")
                    break

                # 打招呼进度
                if progress_callback:
                    pct = int((i + 1) / len(to_greet_list) * 100)
                    progress_callback(pct, f"正在打招呼... {i + 1}/{len(to_greet_list)}")

                if i > 0:
                    if i % GREET_BATCH_SIZE == 0:
                        pause = _human_delay(GREET_BATCH_PAUSE_CENTER, GREET_BATCH_PAUSE_SPREAD)
                        print(f"\n  已连续打招呼 {i} 人，暂停 {int(pause)} 秒降低风控风险...")
                        time.sleep(pause)
                    else:
                        time.sleep(_human_delay(GREET_DELAY_CENTER, GREET_DELAY_SPREAD))

                print(f"  [{i+1}/{len(to_greet_list)}] {candidate['name']} ({candidate['recommend_level']}, {candidate['match_score']}分) {action}...", end=" ")

                # 每 5 个招呼间歇性滚一下，保持虚拟列表持续渲染后续卡片
                if i > 0 and i % 5 == 0:
                    iframe = get_iframe(page)
                    (iframe if iframe else page).run_js('window.scrollBy(0, 400)')
                    time.sleep(_human_delay(0.2, 0.15))

                success, msg = send_greeting_on_list_page(
                    page, candidate['geek_id'], stop_event=stop_event,
                    captcha_callback=captcha_callback)

                if success is None:
                    candidate['greet_sent'] = False
                    candidate.setdefault('followup_status', "未沟通")
                    candidates_all.append(candidate)
                    persist_candidate_greeting_pending(candidate, msg)
                    greet_pending_count += 1
                    consecutive_uncertain += 1
                    print(f"待确认：{msg}")
                    if consecutive_uncertain >= GREET_UNCERTAIN_LIMIT:
                        print(f"\n连续 {consecutive_uncertain} 人发送结果待确认，停止打招呼并请人工核实")
                        break
                    continue
                if success:
                    greet_success_count += 1
                    consecutive_failures = 0  # 重置连续失败计数
                    consecutive_uncertain = 0
                    persist_candidate_greeted(candidate, "auto_list")
                    candidates_all.append(candidate)
                    greeted_in_this_run.append(candidate['geek_id'])
                    print(f"OK")
                else:
                    greet_fail_count += 1
                    consecutive_failures += 1  # 累加连续失败计数
                    consecutive_uncertain = 0
                    candidate['greet_sent'] = False
                    candidate.setdefault('followup_status', "未沟通")
                    candidates_all.append(candidate)
                    print(f"失败：{msg}")

                    # 沟通次数上限是终端条件：达到上限后所有后续打招呼都不会成功
                    if "上限" in msg or "次数" in msg:
                        print(f"\n{'='*60}")
                        print(f"⚠️  今日沟通次数已达上限！")
                        print(f"   BOSS 直聘已弹出升级套餐页面，后续打招呼不会真正发送")
                        print(f"   本次已成功打招呼：{greet_success_count} 人")
                        print(f"{'='*60}")
                        break

                    # 安全验证（滑块/图形验证码）：暂停自动化，等待人工处理
                    if "验证" in msg:
                        print(f"\n{'='*60}")
                        print(f"⚠️  检测到安全验证弹窗！")
                        print(f"   请手动完成验证码后再继续。")
                        print(f"   本次已成功打招呼：{greet_success_count} 人，{greet_fail_count} 人失败")
                        print(f"{'='*60}")
                        break

        except KeyboardInterrupt:
            print(f"\n\n⚠️  检测到中断，保存当前进度...")
            # 中断时立即保存所有数据
            merge_candidates_all(candidates_all)
            if greeted_in_this_run:
                print(f"  本次运行已打招呼 {len(greeted_in_this_run)} 人")
            print(f"✅ 候选人总数：{len(candidates_all)}")
            raise

        print(
            f"\n打招呼完成：成功 {greet_success_count} 人，失败 {greet_fail_count} 人，"
            f"待确认 {greet_pending_count} 人"
        )

    # 保存所有通过的候选人（包含未打招呼的）
    # 用字典索引避免 O(n²) 查找
    existing_index = {c.get('geek_id'): c for c in candidates_all}
    for c in passed_candidates:
        if not c.get('greet_sent'):
            if c.get('geek_id') not in existing_index:
                candidates_all.append(c)
                existing_index[c.get('geek_id')] = c
    merge_candidates_all(candidates_all)

    if stats is not None:
        stats['raw_count'] = len(raw_candidates)
        stats['rule_passed_count'] = rule_pool_count
        stats['passed_count'] = len(passed_candidates)
        stats['greeted_count'] = sum(1 for c in passed_candidates if c.get('greet_sent'))
        if ai_eval:
            stats['ai_eval_count'] = ai_evaluated_count
            stats['ai_downgraded'] = ai_removed_count

    return passed_candidates


def _show_job_navigation_prompt(current_idx, total, next_job_name, confirm_callback=None):
    """多岗位间页面导航提示

    在切换到下一个岗位之前，提示用户手动导航到新岗位的推荐页面。
    - CLI 模式：等待用户按 Enter 确认
    - GUI 模式：通过 confirm_callback 等待用户确认（无倒计时）
    """
    banner = f"""
╔══════════════════════════════════════════════════════════════╗
║  🔄 岗位切换：请导航到下一个岗位的推荐页面                      ║
╠══════════════════════════════════════════════════════════════╣
║  进度：{current_idx}/{total}                                       ║
║  下一个岗位：{next_job_name}
╠══════════════════════════════════════════════════════════════╣
║  BOSS 直聘每个岗位有独立的推荐页面 URL，请手动切换              ║
║  例如：https://www.zhipin.com/web/geek/recommend?jobId=xxxx  ║
╚══════════════════════════════════════════════════════════════╝
"""
    print(banner)

    if confirm_callback:
        # GUI 模式：通过回调等待用户确认（阻塞式，无倒计时）
        confirmed = confirm_callback(current_idx, total, next_job_name)
        if confirmed:
            print("已确认，继续处理...\n")
        else:
            print("用户取消岗位切换\n")
            raise KeyboardInterrupt
    else:
        # CLI 模式：等待 Enter，失败则倒计时
        try:
            input("请导航到新页面后按 Enter 继续...")
            print("已确认，继续处理...\n")
        except (EOFError, OSError):
            # GUI 模式或无终端环境，给用户 15 秒切换页面
            print("(GUI 模式) 请在 15 秒内手动切换到新岗位的推荐页面...")
            for remaining in range(15, 0, -1):
                print(f"  {remaining} 秒后自动继续...", end="\r")
                time.sleep(1)
            print("  继续处理...\n")


def _format_scan_summary(
    status: str,
    total_rule_passed: int,
    total_raw: int,
    total_ai_evaluated: int,
    total_ai_downgraded: int,
    total_passed: int,
    total_greeted: int,
) -> str:
    """Format the final scan summary with rule and AI counts kept separate."""
    prefix = "筛选完成：" if status == "完成" else ""
    message = f"[{status}] {prefix}规则筛选通过 {total_rule_passed}/{total_raw} 人"
    if total_ai_evaluated > 0:
        message += (
            f"，AI复核后淘汰 {total_ai_downgraded} 人，"
            f"最终保留 {total_passed} 人"
        )
    return f"{message}，{total_greeted} 人已打招呼"


def run_smart_scan(args=None, progress_callback=None, confirm_callback=None, stop_event=None, existing_page=None, captcha_callback=None, notice_callback=None, blocking_notice_callback=None):
    """运行智能扫描（支持多岗位）

    参数：
        args: argparse.Namespace 对象，如果为 None 则从命令行解析
        progress_callback: 进度回调 callable(percentage, description)，GUI 模式使用
        confirm_callback: 岗位切换确认回调 callable(current_idx, total, next_job_name) -> bool，GUI 模式使用
        stop_event: threading.Event，设置时立即停止扫描
        existing_page: 已有的浏览器页面对象（GUI 模式传入，避免重复连接）
        captcha_callback: callable(detail) -> bool，检测到验证码时调用，
            用于 GUI 弹窗通知用户。返回 True 继续等待，False 中止。
        notice_callback: callable(title, message)，用于 GUI 展示非阻塞提示。
    """
    import argparse

    # 如果没有传入参数，从命令行解析
    if args is None:
        parser = argparse.ArgumentParser(description='BOSS 直聘候选人智能提取工具')
        parser.add_argument('--clear', action='store_true', help='清空 candidates_all.json 后全新跑')
        parser.add_argument('--keep-greeted', action='store_true', help='清空时保留已打招呼的候选人（仅与 --clear 配合使用）')
        parser.add_argument('--job', type=str, help='指定岗位名称，只跑该岗位')
        parser.add_argument('--greet', action='store_true', help='自动打招呼：对新匹配的候选人自动发送消息')
        parser.add_argument('--re-greet', action='store_true', help='补打招呼：给已匹配但未打招呼的候选人发送消息')
        parser.add_argument('--greet-level', type=str, choices=['strong', 'normal'], default='normal',
                            help='打招呼等级（仅补打招呼模式有效）：strong=仅强烈推荐，normal=强烈推荐 + 推荐（默认）')
        parser.add_argument('--greet-names', type=str, help='点对点打招呼（仅补打招呼模式有效）：指定候选人姓名，多个用逗号分隔')
        parser.add_argument('--list-candidates', action='store_true', help='仅列出候选人，不打招呼')
        parser.add_argument('--rounds', type=int, default=30,
                            help='最大 DOM 滚动轮次（默认 30；深度扫描可手动设为 50-200）')
        parser.add_argument('--max-candidates', type=int, default=API_CANDIDATE_LIMIT_DEFAULT,
                            help=f'API 结构化补全预算（默认 {API_CANDIDATE_LIMIT_DEFAULT}，用于推导最多补全页数）')
        extraction_group = parser.add_mutually_exclusive_group()
        extraction_group.add_argument('--dom-only', action='store_true',
                                      help='仅使用页面滚动 DOM 提取，不进行 listener/API 结构化补全')
        extraction_group.add_argument('--listener-first', action='store_true',
                                      help='使用 DOM 提取 + listener 结构化补全，不进行 API 直调补全')
        parser.add_argument('--verbose', action='store_true', help='输出详细评分信息（显示技能匹配详情）')
        parser.add_argument('--ai-eval', action='store_true', help='启用 AI 辅助评估：对通过筛选的候选人进行 LLM 二次评分')
        args = parser.parse_args()

    # 确定运行模式
    re_greet_mode = args.re_greet
    point_to_point_mode = bool(args.greet_names) and re_greet_mode  # 点对点模式仅在补打招呼时生效

    # 初次扫描时的自动打招呼逻辑（--greet）
    auto_greet_scan = args.greet  # 初次扫描时对所有符合条件的打招呼

    # 补打招呼时的打招呼等级
    greet_levels_allowed = ['强烈推荐', '推荐']  # 默认
    greet_level_text = '强烈推荐 + 推荐'
    if re_greet_mode:
        if args.greet_level == 'strong':
            greet_levels_allowed = ['强烈推荐']
            greet_level_text = '仅强烈推荐'

    # 模式显示
    mode_text = "补打招呼模式" if re_greet_mode else ("全新模式（保留已沟通）" if (args.clear and args.keep_greeted) else ("全新模式" if args.clear else "增量模式"))
    if point_to_point_mode:
        mode_text = f"点对点打招呼模式 ({args.greet_names})"
    greet_text = ""
    if auto_greet_scan:
        greet_level_display = "仅强烈推荐" if args.greet_level == 'strong' else "强烈推荐 + 推荐"
        greet_text = f" + 自动打招呼 ({greet_level_display})"
    elif re_greet_mode:
        greet_text = f" + 打招呼等级 ({greet_level_text})"
    print(f">>> BOSS 直聘候选人智能提取工具 v{_read_app_version()} [{mode_text}{greet_text}]")
    print("="*50)

    # 清空 candidates_all.json（如果指定 --clear）
    if args.clear and os.path.exists(CANDIDATES_PATH):
        candidates_all = load_candidates_all()
        blacklisted = [c for c in candidates_all if c.get('blacklisted')]
        if args.keep_greeted:
            # 保留已打招呼和已屏蔽的候选人
            kept = [c for c in candidates_all if c.get('greet_sent') or c.get('blacklisted')]
            kept_count = len(kept)
            removed = len(candidates_all) - kept_count
            if kept_count > 0:
                save_candidates_all(kept)
            else:
                os.remove(CANDIDATES_PATH)
            print(f"已清空 candidates_all.json（保留 {kept_count} 条已打招呼/黑名单记录，删除 {removed} 条）")
        else:
            removed = len(candidates_all) - len(blacklisted)
            if blacklisted:
                save_candidates_all(blacklisted)
                print(f"已清空 candidates_all.json（保留 {len(blacklisted)} 条黑名单记录，删除 {removed} 条）")
            else:
                os.remove(CANDIDATES_PATH)
                print("已清空 candidates_all.json")

    # 补打招呼模式：直接处理，不需要打开浏览器扫描
    if re_greet_mode:
        print(f"\n[补打招呼模式] 读取 candidates_all.json，给未打招呼的候选人发送消息...")
        print("="*50)

        # 读取已有数据
        candidates_all = load_candidates_all()
        if not candidates_all:
            print("candidates_all.json 为空或不存在")
        else:
            # 解析点对点打招呼的姓名列表
            greet_names_list = None
            if args.greet_names:
                greet_names_list = [name.strip() for name in args.greet_names.split(',')]
                print(f"点对点打招呼：{greet_names_list}")

            # 筛选需要打招呼的候选人
            if greet_names_list:
                # 点对点模式：只筛选指定姓名的候选人
                to_greet = [c for c in candidates_all
                           if c.get('name') in greet_names_list
                           and not c.get('greet_sent', False)
                           and not c.get('blacklisted')]
                filter_text = f" (姓名匹配：{greet_names_list})"
            else:
                # 自动模式：根据推荐等级筛选
                to_greet = [c for c in candidates_all
                           if c.get('recommend_level') in greet_levels_allowed
                           and not c.get('greet_sent', False)
                           and not c.get('greet_confirmation_pending')
                           and not c.get('blacklisted')
                           and not c.get('manual_review_required')]
                blocked_count = sum(
                    1 for c in candidates_all
                    if c.get('recommend_level') in greet_levels_allowed
                    and not c.get('greet_sent', False)
                    and c.get('manual_review_required')
                )
                if blocked_count:
                    print(f"已跳过 {blocked_count} 人：需要人工确认后再打招呼")
                filter_text = f" ({greet_level_text})"

            if not to_greet:
                print(f"没有需要补打招呼的候选人{filter_text}")
            else:
                print(f"找到 {len(to_greet)} 个需要补打招呼的候选人{filter_text}:")
                for c in to_greet[:10]:
                    print(f"  - {c.get('name')} ({c.get('recommend_level')}, {c.get('match_score')}分)")
                if len(to_greet) > 10:
                    print(f"  ... 还有 {len(to_greet) - 10} 个")
                if len(to_greet) > AUTO_GREET_RUN_LIMIT:
                    remaining_count = len(to_greet) - AUTO_GREET_RUN_LIMIT
                    limit_msg = (
                        f"为降低 BOSS 风控风险，本轮最多补打招呼 {AUTO_GREET_RUN_LIMIT} 人，"
                        f"剩余 {remaining_count} 人下次继续。\n\n"
                        "下次再次运行补打招呼即可：程序会跳过已经标记为已打招呼的人，"
                        "未打招呼的候选人会继续进入补打招呼队列。"
                    )
                    print(f"  {limit_msg.replace(chr(10), ' ')}")
                    if notice_callback:
                        try:
                            notice_callback("已达到本轮补打招呼上限", limit_msg)
                        except Exception:
                            pass
                    to_greet = to_greet[:AUTO_GREET_RUN_LIMIT]

                # 需要打开浏览器进行打招呼
                if existing_page:
                    page = existing_page
                    print("使用 GUI 已连接的浏览器")
                else:
                    page = ChromiumPage()
                    print("\n浏览器已打开，请手动导航到候选人推荐页面")
                print("等待 3 秒...")
                time.sleep(_human_delay(3.0, 2.0))

                # 执行补打招呼
                success_count = 0
                fail_count = 0
                skip_count = 0
                consecutive_uncertain = 0

                try:
                    for i, c in enumerate(to_greet):
                        if stop_event and stop_event.is_set():
                            print(f"\n⏹ 用户停止，已跳过剩余 {len(to_greet) - i} 人")
                            break
                        geek_id = c.get('geek_id')
                        name = c.get('name', '未知')
                        if i > 0:
                            if i % GREET_BATCH_SIZE == 0:
                                pause = _human_delay(GREET_BATCH_PAUSE_CENTER, GREET_BATCH_PAUSE_SPREAD)
                                print(f"\n已连续打招呼 {i} 人，暂停 {int(pause)} 秒降低风控风险...")
                                time.sleep(pause)
                            else:
                                time.sleep(_human_delay(GREET_DELAY_CENTER, GREET_DELAY_SPREAD))
                        print(f"[{i+1}/{len(to_greet)}] 正在向 {name} 打招呼...", end=" ")
                        success, msg = send_greeting_on_list_page(page, geek_id, stop_event=stop_event, captcha_callback=captcha_callback)
                        if success is None:
                            skip_count += 1
                            persist_candidate_greeting_pending(c, msg)
                            consecutive_uncertain += 1
                            print(f"待确认：{msg}")
                            if consecutive_uncertain >= GREET_UNCERTAIN_LIMIT:
                                print(f"\n连续 {consecutive_uncertain} 人发送结果待确认，停止补打招呼并请人工核实")
                                break
                            continue
                        if success:
                            success_count += 1
                            consecutive_uncertain = 0
                            persist_candidate_greeted(c, "regreet_list")
                            print("OK")
                        else:
                            fail_count += 1
                            consecutive_uncertain = 0
                            print(f"失败：{msg}")
                            # 沟通次数上限是终端条件
                            if "上限" in msg or "次数" in msg:
                                print(f"\n{'='*60}")
                                print(f"⚠️  今日沟通次数已达上限！")
                                print(f"   BOSS 直聘已弹出升级套餐页面，后续打招呼不会真正发送")
                                print(f"   本次已成功打招呼：{success_count} 人")
                                print(f"{'='*60}")
                                break

                except KeyboardInterrupt:
                    print(f"\n\n检测到中断，保存当前进度...")
                    merge_candidates_all(candidates_all)
                    # 生成 Excel 文件
                    if export_to_excel(candidates_all, CANDIDATES_XLSX_PATH):
                        print(f"[SAVE] Excel 文件：{CANDIDATES_XLSX_PATH.name}")
                    print(f"已保存 {success_count} 个成功打招呼的候选人状态")
                    raise

                print(f"\n补打招呼完成：成功 {success_count} 人，失败 {fail_count} 人，待确认 {skip_count} 人")
                merge_candidates_all(candidates_all)
                print(f"已更新 candidates_all.json")

        print("\n--- 浏览器保持打开 ---")
        return  # 补打招呼模式结束，直接返回

    page = None
    all_candidates = []  # 保存所有岗位的候选人

    # 加载 AI 评估所需的 API 配置
    ai_api_config = getattr(args, 'api_config', None)
    ai_api_key = getattr(args, 'api_key', None)
    if getattr(args, 'ai_eval', False) and (ai_api_config is None or ai_api_key is None):
        try:
            with open('api_config.json', 'r', encoding='utf-8') as f:
                ai_api_config = json.load(f)
            from security import get_api_key
            ai_api_key = get_api_key(ai_api_config.get('api_provider', ''), ai_api_config.get('base_url', ''))
            if not ai_api_key:
                print("⚠️  AI 评估需要 API Key，但未配置，将跳过 AI 评估")
                args.ai_eval = False
            else:
                model_name = ai_api_config.get('model', 'unknown')
                print(f"AI 辅助评估已启用（模型：{model_name}）")
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️  加载 API 配置失败：{e}，将跳过 AI 评估")
            args.ai_eval = False
    elif getattr(args, 'ai_eval', False):
        model_name = ai_api_config.get('model', 'unknown') if ai_api_config else 'unknown'
        print(f"AI 辅助评估已启用（模型：{model_name}）")

    # 统计累计（提前初始化，防止异常路径上 NameError）
    job_stats = {}
    total_raw = total_rule_passed = total_passed = 0
    total_greeted = total_ai_evaluated = total_ai_downgraded = 0

    try:
        if existing_page:
            page = existing_page
            print("使用 GUI 已连接的浏览器")
        else:
            print("正在连接到浏览器...")
            page = ChromiumPage()
            print("\n浏览器已打开，请手动导航到候选人推荐页面")
            print("例如：https://www.zhipin.com/web/chat/recommend")
        print("请确保页面完全加载后，等待 3 秒...")
        time.sleep(_human_delay(3.0, 2.0))

        job_rules, default_rule = load_job_config()

        # 确定要运行的岗位列表
        jobs_to_run = []
        if args.job:
            resolved_job = _resolve_job_name(args.job, job_rules)
            if resolved_job:
                jobs_to_run = [resolved_job]
            else:
                print(f"错误：未找到岗位 '{args.job}'")
                print(f"可用岗位：{', '.join(job_rules.keys())}")
                return
        else:
            jobs_to_run = list(job_rules.keys())

        print(f"\n将要运行 {len(jobs_to_run)} 个岗位：{', '.join(jobs_to_run)}")
        if default_rule:
            print(f"(另有 default 默认规则，不作为岗位运行)")
        print("="*50)

        # 逐个岗位处理
        for idx, job_name in enumerate(jobs_to_run, 1):
            # 多岗位间页面导航提示（第一个岗位之前不需要）
            if idx > 1:
                _show_job_navigation_prompt(idx, len(jobs_to_run), job_name, confirm_callback=confirm_callback)

            print(f"\n{'='*50}")
            print(f">>> 处理岗位 {idx}/{len(jobs_to_run)}: {job_name}")
            print(f"{'='*50}")

            rule = job_rules[job_name]
            job_info = {
                "job_id": "unknown",
                "job_name": job_name,
                "rule": rule,
                "rule_key": job_name
            }

            print(f"过滤规则：经验≥{rule.get('min_exp', 0)}年，学历≥{rule.get('edu', '不限')}")
            print(f"Keywords: {[k.get('name', k) if isinstance(k, dict) else k for k in rule.get('keywords', [])][:5]}...")

            if getattr(args, 'dom_only', False):
                extraction_mode = "dom"
            elif getattr(args, 'listener_first', False):
                extraction_mode = "listener"
            else:
                extraction_mode = "api"

            candidates = smart_scan_candidates(page, job_info, auto_greet=auto_greet_scan,
                                               max_rounds=args.rounds, verbose=args.verbose,
                                               greet_level=args.greet_level, greet_names_list=None,
                                               list_candidates=args.list_candidates,
                                               progress_callback=progress_callback,
                                               stop_event=stop_event,
                                               ai_eval=getattr(args, 'ai_eval', False),
                                               api_config=ai_api_config,
                                               api_key=ai_api_key,
                                               captcha_callback=captcha_callback,
                                               notice_callback=notice_callback,
                                               blocking_notice_callback=blocking_notice_callback,
                                               max_candidates=getattr(args, 'max_candidates', API_CANDIDATE_LIMIT_DEFAULT),
                                               extraction_mode=extraction_mode,
                                               stats=job_stats)
            all_candidates.extend(candidates)
            total_raw += job_stats.get('raw_count', 0)
            total_rule_passed += job_stats.get(
                'rule_passed_count', job_stats.get('passed_count', 0)
            )
            total_passed += job_stats.get('passed_count', 0)
            total_greeted += job_stats.get('greeted_count', 0)
            total_ai_evaluated += job_stats.get('ai_eval_count', 0)
            total_ai_downgraded += job_stats.get('ai_downgraded', 0)

        # 最后生成 Excel 文件
        existing_all = load_candidates_all()
        excel_file = CANDIDATES_XLSX_PATH
        if export_to_excel(existing_all, excel_file):
            print(f"[SAVE] Excel 文件：{excel_file.name}")
        else:
            print("[WARN] Excel 导出失败")

        # 全部岗位处理完毕，更新进度为最终状态
        if progress_callback:
            msg = _format_scan_summary(
                "完成",
                total_rule_passed,
                total_raw,
                total_ai_evaluated,
                total_ai_downgraded,
                total_passed,
                total_greeted,
            )
            progress_callback(100, msg)

    except StopRequested:
        print(f"\n\n⏹ 用户停止，保存当前进度...")
        _save_progress_on_exit()
        if progress_callback:
            stop_msg = _format_scan_summary(
                "已停止",
                total_rule_passed,
                total_raw,
                total_ai_evaluated,
                total_ai_downgraded,
                total_passed,
                total_greeted,
            )
            progress_callback(100, stop_msg)

    except KeyboardInterrupt:
        print(f"\n\n检测到中断，保存当前进度...")
        _save_progress_on_exit()
        if progress_callback:
            stop_msg = _format_scan_summary(
                "已停止",
                total_rule_passed,
                total_raw,
                total_ai_evaluated,
                total_ai_downgraded,
                total_passed,
                total_greeted,
            )
            progress_callback(100, stop_msg)
        raise

    except Exception as e:
        print(f"程序执行出错：{e}")
        import traceback
        print(traceback.format_exc())
        try:
            _save_progress_on_exit()
        except Exception as save_err:
            print(f"保存进度时出错：{save_err}")
        if progress_callback:
            progress_callback(100, f"[出错] {str(e)[:30]}")

    finally:
        if page:
            print("\n--- 浏览器保持打开 ---")


if __name__ == "__main__":
    run_smart_scan()
